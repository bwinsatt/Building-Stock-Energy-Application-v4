"""
Generate Emissions_and_Constants.xlsx workbook with 4 sheets:
  1. Energy Conversions
  2. Fossil Fuel Emission Factors
  3. eGRID Electricity Emission Factors (from zipcode_lookup.json)
  4. Model Accuracy
"""

import json
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
ZIPCODE_LOOKUP = PROJECT_ROOT / "backend" / "app" / "data" / "zipcode_lookup.json"
OUTPUT_PATH = Path(__file__).resolve().parent / "Emissions_and_Constants.xlsx"

# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, num_cols):
    """Apply header formatting to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_fit_columns(ws, max_width=60):
    """Approximate auto-fit by scanning cell values."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def apply_alternating_rows(ws, start_row=2):
    """White / #F2F3F4 alternating fills starting from start_row."""
    for row_idx in range(start_row, ws.max_row + 1):
        if (row_idx - start_row) % 2 == 1:
            for cell in ws[row_idx]:
                cell.fill = ALT_ROW_FILL


def finalize_sheet(ws, num_cols):
    """Freeze top row, add autofilter, auto-fit columns."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"
    auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 1 – Energy Conversions
# ---------------------------------------------------------------------------
def build_energy_conversions(wb):
    ws = wb.active
    ws.title = "Energy Conversions"

    headers = ["Constant Name", "Value", "Unit", "Description"]
    ws.append(headers)

    data = [
        ("KWH_TO_KBTU", 3.412, "kBtu/kWh", "1 kWh = 3.412 kBtu"),
        ("KWH_TO_THERMS", 0.03412, "therms/kWh", "1 kWh = 0.03412 therms"),
        ("KWH_PER_GALLON_FUEL_OIL", 40.6, "kWh/gallon", "1 gallon #2 fuel oil = 40.6 kWh"),
        ("KWH_PER_GALLON_PROPANE", 26.8, "kWh/gallon", "1 gallon propane = 26.8 kWh"),
        ("KWH_PER_THERM", 29.31, "kWh/therm", "1 therm = 29.31 kWh"),
        ("TONS_TO_KBTUH", 12.0, "kBtu/h per ton", "1 ton cooling capacity = 12 kBtu/h"),
        ("SQFT_PER_M2", 10.7639, "ft\u00b2/m\u00b2", "1 m\u00b2 = 10.7639 ft\u00b2"),
        ("STORY_HEIGHT_COMMERCIAL", 12, "feet", "Default commercial floor-to-floor height"),
        ("STORY_HEIGHT_RESIDENTIAL", 10, "feet", "Default residential floor-to-floor height"),
        ("ASPECT_RATIO", 1.8, "dimensionless", "Default building length/width ratio for geometric sizing"),
    ]

    for row in data:
        ws.append(row)

    # Number formatting for the Value column
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=2)
        if isinstance(cell.value, float):
            # Use up to 5 decimal places
            cell.number_format = "0.#####"

    style_header(ws, len(headers))
    apply_alternating_rows(ws)
    finalize_sheet(ws, len(headers))


# ---------------------------------------------------------------------------
# Sheet 2 – Fossil Fuel Emission Factors
# ---------------------------------------------------------------------------
def build_fossil_fuel_factors(wb):
    ws = wb.create_sheet("Fossil Fuel Emission Factors")

    headers = ["Fuel Type", "Emission Factor (kg CO2e/kBtu)", "Source", "Notes"]
    ws.append(headers)

    data = [
        ("Natural Gas", 0.05311, "EPA GHG Emission Factors Hub", "Fixed national factor"),
        ("Propane", 0.06195, "EPA GHG Emission Factors Hub", "Fixed national factor"),
        (
            "Fuel Oil (#2)",
            0.07421,
            "EPA GHG Emission Factors Hub",
            "Fixed national factor; highest carbon intensity",
        ),
        (
            "District Heating",
            0.06640,
            "EPA GHG Emission Factors Hub",
            "Approximation; varies by actual district system fuel mix",
        ),
    ]

    for row in data:
        ws.append(row)

    # 5 decimal places for emission factor column
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).number_format = "0.00000"

    style_header(ws, len(headers))
    apply_alternating_rows(ws)
    finalize_sheet(ws, len(headers))


# ---------------------------------------------------------------------------
# Sheet 3 – eGRID Electricity Emission Factors
# ---------------------------------------------------------------------------
def build_egrid_factors(wb):
    ws = wb.create_sheet("eGRID Electricity Factors")

    # Top note
    note = (
        "Electricity emission factors are region-specific from EPA eGRID. "
        "They vary significantly based on the local grid's fuel mix "
        "(coal, gas, nuclear, renewables). Buildings in cleaner grid regions "
        "get less emissions reduction credit from electricity savings."
    )
    ws.append([note])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    note_cell = ws.cell(row=1, column=1)
    note_cell.font = Font(italic=True)
    note_cell.alignment = Alignment(wrap_text=True)

    # Blank separator row
    ws.append([])

    # Headers in row 3
    headers = [
        "State",
        "eGRID Subregion",
        "Electricity Emission Factor (kg CO2e/kWh)",
        "Emission Factor (lb CO2e/MWh)",
        "Climate Zone (example)",
        "HDD65F (example)",
        "CDD65F (example)",
    ]
    ws.append(headers)
    header_row = 3

    # Read zipcode lookup
    with open(ZIPCODE_LOOKUP, "r") as f:
        lookup = json.load(f)

    # Deduplicate by (state, egrid_subregion) — first occurrence wins
    # Some entries lack egrid_subregion / emission factor; skip those.
    seen = set()
    rows = []
    for _prefix, entry in lookup["prefixes"].items():
        if "egrid_subregion" not in entry:
            continue
        key = (entry["state"], entry["egrid_subregion"])
        if key not in seen:
            seen.add(key)
            ef = entry["electricity_emission_factor_kg_co2e_per_kwh"]
            rows.append(
                (
                    entry["state"],
                    entry["egrid_subregion"],
                    ef,
                    ef * 2204.62,  # kg/kWh -> lb/MWh  (×1000 kWh/MWh × 2.20462 lb/kg = ×2204.62)
                    entry["climate_zone"],
                    entry["hdd65f"],
                    entry["cdd65f"],
                )
            )

    # Sort by state then subregion
    rows.sort(key=lambda r: (r[0], r[1]))

    for row in rows:
        ws.append(row)

    # Format header row (row 3)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Number formats for data rows
    data_start = header_row + 1
    for row_idx in range(data_start, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).number_format = "0.00000"   # kg CO2e/kWh
        ws.cell(row=row_idx, column=4).number_format = "#,##0.00"  # lb CO2e/MWh
        ws.cell(row=row_idx, column=6).number_format = "#,##0.0"   # HDD
        ws.cell(row=row_idx, column=7).number_format = "#,##0.0"   # CDD

    # Alternating rows starting after header
    for row_idx in range(data_start, ws.max_row + 1):
        if (row_idx - data_start) % 2 == 1:
            for cell in ws[row_idx]:
                cell.fill = ALT_ROW_FILL

    # Freeze pane below header, autofilter on header row
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"
    )
    auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 4 – Model Accuracy
# ---------------------------------------------------------------------------
def build_model_accuracy(wb):
    ws = wb.create_sheet("Model Accuracy")

    headers = ["Dataset", "Fuel Type", "Mean MAE (kBtu/ft\u00b2)", "Mean R\u00b2", "Avg Training Samples", "Notes"]
    ws.append(headers)

    data = [
        ("ComStock", "Electricity", 12.04, 0.817, "~18,000-40,000", "Highest absolute error due to diverse end uses"),
        ("ComStock", "Natural Gas", 9.93, 0.662, "~18,000-40,000", "Lower R\u00b2 reflects gas heating variability"),
        ("ComStock", "Fuel Oil", 0.60, 0.832, "~18,000-40,000", "Low MAE; few buildings use fuel oil"),
        ("ComStock", "Propane", 0.72, 0.891, "~18,000-40,000", "Highest R\u00b2 among fuels"),
        ("ComStock", "District Heating", 0.24, 0.836, "~18,000-40,000", "Very low MAE; limited to urban buildings"),
        ("ResStock", "Electricity", 1.45, 0.733, "~67,000-75,000", "Lower absolute values than commercial"),
        ("ResStock", "Natural Gas", 0.58, 0.761, "~67,000-75,000", "Good performance for residential gas"),
        ("ResStock", "Fuel Oil", 0.06, 0.918, "~67,000-75,000", "Highest R\u00b2; limited fuel oil usage"),
        ("ResStock", "Propane", 0.02, 0.796, "~67,000-75,000", "Very low MAE; few propane buildings"),
    ]

    for row in data:
        ws.append(row)

    # Number formatting
    for row_idx in range(2, 2 + len(data)):
        ws.cell(row=row_idx, column=3).number_format = "0.00"   # MAE
        ws.cell(row=row_idx, column=4).number_format = "0.000"  # R²

    # Footer notes
    notes = [
        "MAE = Mean Absolute Error on held-out test set (20% split, random state 42)",
        "R\u00b2 = Coefficient of determination (1.0 = perfect prediction)",
        "These are averages across all upgrade models per fuel type",
        "Baseline model (upgrade 0) electricity: ComStock R\u00b2=0.844, MAE=13.36 kBtu/ft\u00b2",
        "Minor fuels (oil, propane, district) show higher R\u00b2 due to less consumption variability",
    ]
    ws.append([])  # blank row
    for note in notes:
        ws.append([note])
        row_idx = ws.max_row
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        ws.cell(row=row_idx, column=1).font = Font(italic=True, size=9)

    style_header(ws, len(headers))
    apply_alternating_rows(ws, start_row=2)
    finalize_sheet(ws, len(headers))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()

    build_energy_conversions(wb)
    build_fossil_fuel_factors(wb)
    build_egrid_factors(wb)
    build_model_accuracy(wb)

    wb.save(str(OUTPUT_PATH))
    print(f"Workbook saved to {OUTPUT_PATH}")
    print(f"File size: {os.path.getsize(OUTPUT_PATH):,} bytes")

    # Quick summary
    wb2 = __import__("openpyxl").load_workbook(str(OUTPUT_PATH))
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  Sheet '{name}': {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
