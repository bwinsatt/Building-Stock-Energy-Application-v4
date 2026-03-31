"""
Generate Applicability_Rules.xlsx summarizing upgrade applicability rules
for ComStock and ResStock datasets.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True)
ALT_ROW_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
PACKAGE_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")

# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

COMSTOCK_HEADERS = [
    "Upgrade ID",
    "Upgrade Name",
    "Category",
    "Applicable To (conditions)",
    "Not Applicable If",
    "Features Checked",
    "Is Package",
]

COMSTOCK_DATA = [
    [1, "Variable Speed HP RTU, Electric Backup", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [2, "Variable Speed HP RTU, Original Heating Fuel Backup", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [3, "Variable Speed HP RTU, Electric Backup, Energy Recovery", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [4, "Standard Performance HP RTU, Electric Backup", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [5, "Standard Performance HP RTU using Lab Data, Electric Backup", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [6, "Standard Performance HP RTU, Electric Backup + Roof Insulation", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [7, "Standard Performance HP RTU, Electric Backup + New Windows", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [8, "Standard Performance HP RTU, Electric Backup, 32F Min Compressor Lockout", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [9, "Standard Performance HP RTU, Electric Backup, 2F Unoccupied Htg Setback", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [10, "Commercial HVAC Challenge HP RTU, Electric Backup", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump (ASHP/GSHP/WSHP)",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [11, "Advanced RTU", "HVAC",
     "Small Packaged Unit HVAC buildings",
     "DOAS ventilation OR already has heat pump",
     "in.hvac_category, in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [12, "VRF with DOAS", "HVAC",
     "Most building types",
     "DOAS ventilation OR District cooling OR already has heat pump OR Food Service/Healthcare",
     "in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type, in.comstock_building_type_group",
     "No"],
    [13, "VRF with 25pct Upsizing Allowance", "HVAC",
     "Most building types",
     "DOAS ventilation OR District cooling OR already has heat pump OR Food Service/Healthcare",
     "in.hvac_vent_type, in.hvac_cool_type, in.hvac_heat_type, in.comstock_building_type_group",
     "No"],
    [14, "DOAS HP Minisplits", "HVAC",
     "Small Packaged Unit or Residential Style Central Systems",
     "Already has heat pump cooling/heating",
     "in.hvac_category, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [15, "HP Boiler, Electric Backup", "HVAC",
     "Buildings with boiler",
     "No boiler present",
     "in.hvac_heat_type",
     "No"],
    [16, "HP Boiler, Gas Backup", "HVAC",
     "Buildings with gas boiler",
     "No boiler OR heating fuel not Natural Gas",
     "in.hvac_heat_type, in.heating_fuel",
     "No"],
    [17, "Condensing Gas Boilers", "HVAC",
     "Buildings with gas boiler",
     "No boiler OR heating fuel not Natural Gas",
     "in.hvac_heat_type, in.heating_fuel",
     "No"],
    [18, "Electric Resistance Boilers", "HVAC",
     "Buildings with boiler",
     "No boiler present",
     "in.hvac_heat_type",
     "No"],
    [19, "Air Side Economizers for AHUs", "HVAC Controls",
     "Most HVAC configurations",
     "DOAS or Zone terminal ventilation",
     "in.hvac_vent_type",
     "No"],
    [20, "Demand Control Ventilation", "HVAC Controls",
     "Most building types",
     "DOAS/Zone terminal ventilation OR Lodging/Food Service",
     "in.hvac_vent_type, in.comstock_building_type_group",
     "No"],
    [21, "Energy Recovery for AHUs", "HVAC Controls",
     "Most building types",
     "Zone terminal ventilation OR Food Service",
     "in.hvac_vent_type, in.comstock_building_type_group",
     "No"],
    [22, "Advanced RTU Controls", "HVAC Controls",
     "Small Packaged Unit buildings",
     "Not Small Packaged Unit",
     "in.hvac_category",
     "No"],
    [23, "Unoccupied AHU Control", "HVAC Controls",
     "Most HVAC configurations",
     "DOAS or Zone terminal ventilation",
     "in.hvac_vent_type",
     "No"],
    [24, "Fan Static Pressure Reset for Multizone VAV", "HVAC Controls",
     "Multizone CAV/VAV buildings",
     "Not Multizone CAV/VAV",
     "in.hvac_category",
     "No"],
    [25, "Thermostat Setbacks", "HVAC Controls",
     "All buildings", "(none)", "(none)", "No"],
    [26, "VFD Pumps", "HVAC Controls",
     "Buildings with boiler OR WCC/District cooling",
     "No boiler AND no WCC/District cooling",
     "in.hvac_heat_type, in.hvac_cool_type",
     "No"],
    [27, "Ideal Thermal Air Loads", "Benchmark",
     "All buildings", "(none)", "(none)", "No"],
    [28, "Hydronic Water-to-Water Geothermal Heat Pump", "HVAC",
     "Buildings with boiler AND WCC/ACC cooling",
     "District heating/cooling OR already has GSHP/WSHP",
     "in.hvac_heat_type, in.hvac_cool_type",
     "No"],
    [29, "Packaged Water-to-Air Geothermal Heat Pump", "HVAC",
     "Small Packaged Unit or Residential Central",
     "District heating/cooling OR already has GSHP/WSHP",
     "in.hvac_category, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [30, "Console Water-to-Air Geothermal Heat Pump", "HVAC",
     "Zone-by-Zone HVAC",
     "District heating/cooling OR already has GSHP/WSHP",
     "in.hvac_category, in.hvac_cool_type, in.hvac_heat_type",
     "No"],
    [31, "Chiller Replacement", "HVAC",
     "Buildings with WCC or ACC chiller",
     "No chiller (not WCC or ACC)",
     "in.hvac_cool_type",
     "No"],
    [32, "DF: Thermostat Load Shed, Daily Bldg Peak Reduction", "Demand Flex",
     "Office, Warehouse, Education only",
     "Other building types",
     "in.comstock_building_type_group",
     "No"],
    [33, "DF: Thermostat Load Shift, Daily Bldg Peak Reduction", "Demand Flex",
     "Office, Warehouse, Education only",
     "Other building types",
     "in.comstock_building_type_group",
     "No"],
    [34, "DF: Lighting Load Shed, Daily Bldg Peak Reduction", "Demand Flex",
     "Office, Warehouse, Education only",
     "Other building types",
     "in.comstock_building_type_group",
     "No"],
    [35, "DF: Lighting Load Shed, Daily GHG Emission Reduction", "Demand Flex",
     "Office, Warehouse, Education only",
     "Other building types",
     "in.comstock_building_type_group",
     "No"],
    [36, "DF: Thermostat Load Shed, Grid Peak Reduction", "Demand Flex",
     "Office, Warehouse, Education only",
     "Other building types",
     "in.comstock_building_type_group",
     "No"],
    [37, "DF: Lighting Load Shed, Grid Peak Reduction", "Demand Flex",
     "Office, Warehouse, Education only",
     "Other building types",
     "in.comstock_building_type_group",
     "No"],
    [38, "DF: Plug Load Control (GEB Gem), Grid Peak Reduction", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [39, "DF: Lighting Control (GEB Gem), Grid Peak Reduction", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [40, "DF: Thermostat Control (GEB Gem), Grid Peak Reduction", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [41, "DF: Preheat (GEB Gem), Grid Peak Reduction", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [42, "DF: Precool (GEB Gem), Grid Peak Reduction", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [43, "LED Lighting", "Lighting",
     "Buildings without LED lighting",
     "Already has gen4_led or gen5_led",
     "in.interior_lighting_generation",
     "No"],
    [44, "Lighting Controls", "Lighting",
     "All buildings", "(none)", "(none)", "No"],
    [45, "Electric Kitchen Equipment", "Other",
     "Food Sales or Food Service",
     "Other building types",
     "in.comstock_building_type_group",
     "No"],
    [46, "Photovoltaics with 40pct Roof Coverage", "Renewable",
     "All buildings", "(none)", "(none)", "No"],
    [47, "Photovoltaics with 40pct Roof Coverage + Battery Storage", "Renewable",
     "All buildings", "(none)", "(none)", "No"],
    [48, "Wall Insulation", "Envelope",
     "Most wall types",
     "Metal Building wall construction",
     "in.wall_construction_type",
     "No"],
    [49, "Roof Insulation", "Envelope",
     "All buildings", "(none)", "(none)", "No"],
    [50, "Secondary Windows", "Envelope",
     "Buildings without triple-pane",
     "Already has Triple window type",
     "in.window_type",
     "No"],
    [51, "Window Film", "Envelope",
     "All buildings", "(none)", "(none)", "No"],
    [52, "New Windows", "Envelope",
     "All buildings", "(none)", "(none)", "No"],
    [53, "Code Envelope", "Envelope",
     "All buildings", "(none)", "(none)", "No"],
    [54, "Package 1: Wall + Roof Insulation + New Windows", "Envelope Package",
     "Most wall types",
     "Metal Building wall construction",
     "in.wall_construction_type",
     "Yes (48, 49, 52)"],
    [55, "Package 2: LED Lighting + Variable Speed HP RTU or HP Boilers", "HVAC+Lighting Package",
     "Small Packaged Unit OR has boiler, without LED",
     "Already LED AND (not Small Packaged AND no boiler)",
     "in.interior_lighting_generation, in.hvac_category, in.hvac_heat_type",
     "Yes (43, 1, 15)"],
    [56, "Package 3: Package 2 with Standard Performance HP RTU", "HVAC+Lighting Package",
     "Small Packaged Unit OR has boiler, without LED",
     "Already LED",
     "in.interior_lighting_generation, in.hvac_category",
     "Yes (43, 4)"],
    [57, "Package 4: Package 1 + Package 2", "Comprehensive Package",
     "Combines Package 1 + Package 2 rules",
     "Metal Building wall OR already LED",
     "Multiple",
     "Yes (48, 49, 52, 43, 1, 15)"],
    [58, "Package 5: VS HP RTU or HP Boilers + Economizer + DCV + Energy Recovery", "HVAC Package",
     "Small Packaged Unit, not DOAS, no HP",
     "DOAS or existing HP",
     "Multiple",
     "Yes (1, 15, 19, 20, 21)"],
    [59, "Package 6: Hydronic GHP or Packaged GHP or Console GHP", "HVAC Package",
     "At least one of upgrades 28/29/30 applies",
     "Already has GSHP/WSHP",
     "Multiple",
     "Yes (28, 29, 30)"],
    [60, "Package 7: DF Lighting + Thermostat Load Shed, Grid Peak + PV", "Demand Flex Package",
     "All buildings", "(none)", "(none)", "Yes"],
    [61, "Package 8: DF Lighting + Thermostat Load Shed, Daily Bldg Peak + PV", "Demand Flex Package",
     "All buildings", "(none)", "(none)", "Yes"],
    [62, "Package 9: DF Lighting + Thermostat Load Shed, Daily Bldg Peak", "Demand Flex Package",
     "All buildings", "(none)", "(none)", "Yes"],
    [63, "Package 10: Package 1 + Package 6", "Comprehensive Package",
     "Not Metal Building wall AND Package 6 applies",
     "Metal Building wall OR GHP not applicable",
     "in.wall_construction_type + GHP rules",
     "Yes (48, 49, 52, 28)"],
    [64, "Package 11: Package 10 + LED Lighting", "Comprehensive Package",
     "Not Metal Building, no LED, AND Package 6 applies",
     "Metal Building OR LED OR GHP not applicable",
     "Multiple",
     "Yes (48, 49, 52, 28, 43)"],
    [65, "Package 12: PV with 40pct Roof Coverage + Roof Insulation", "Envelope+Renewable Package",
     "All buildings", "(none)", "(none)", "Yes (46, 49)"],
]

RESSTOCK_HEADERS = COMSTOCK_HEADERS  # Same column structure

RESSTOCK_DATA = [
    [1, "Natural Gas Furnace 95% AFUE for All Dwellings with Ducts", "HVAC",
     "Natural Gas + Ducted Heating",
     "Other fuel or non-ducted",
     "in.heating_fuel, in.hvac_heating_type",
     "No"],
    [2, "Propane or Fuel Oil Furnace 95% AFUE", "HVAC",
     "Propane/Fuel Oil + Ducted Heating",
     "Other fuel or non-ducted",
     "in.heating_fuel, in.hvac_heating_type",
     "No"],
    [3, "Minimum Efficiency Boilers, Furnaces, and ACs Circa 2025", "HVAC",
     "All buildings", "(none)", "(none)", "No"],
    [4, "Typical Cold Climate Ducted ASHP with Detailed Performance Data", "HVAC",
     "Buildings without heat pump",
     "Already has heat pump",
     "in.hvac_heating_type",
     "No"],
    [5, "Dual Fuel Heating System", "HVAC",
     "Non-electric, non-HP buildings",
     "Already has heat pump OR electric heating",
     "in.hvac_heating_type, in.heating_fuel",
     "No"],
    [6, "Single-Speed GHP with Thermally Enhanced Grout and Pipes", "HVAC",
     "Ducted Heating, not heat pump",
     "Non-ducted OR already heat pump",
     "in.hvac_heating_type",
     "No"],
    [7, "Dual-Speed GHP with Thermally Enhanced Grout and Pipes", "HVAC",
     "Ducted Heating, not heat pump",
     "Non-ducted OR already heat pump",
     "in.hvac_heating_type",
     "No"],
    [8, "Variable-Speed GHP with Thermally Enhanced Grout and Pipes", "HVAC",
     "Ducted Heating, not heat pump",
     "Non-ducted OR already heat pump",
     "in.hvac_heating_type",
     "No"],
    [9, "Heat Pump Water Heater", "Water Heating",
     "Most water heater fuels",
     'Water heater fuel is "Other Fuel"',
     "in.water_heater_fuel",
     "No"],
    [10, "High Efficiency Natural Gas Tankless Water Heater", "Water Heating",
     "Natural Gas water heater",
     "Other water heater fuel",
     "in.water_heater_fuel",
     "No"],
    [11, "Air Sealing", "Envelope",
     "All buildings", "(none)", "(none)", "No"],
    [12, "Attic Floor Insulation for Unfinished Attics", "Envelope",
     "All buildings", "(none)", "(none)", "No"],
    [13, "Duct Sealing and Insulation", "Envelope",
     "Ducted heating buildings",
     "Non-ducted heating",
     "in.hvac_heating_type",
     "No"],
    [14, "Drill and Fill Wall Insulation with Air Sealing", "Envelope",
     "Wood Frame walls",
     "Non-wood frame walls",
     "in.geometry_wall_type",
     "No"],
    [15, "Air Sealing + Attic Floor Insulation + Duct Sealing", "Envelope",
     "Ducted heating buildings",
     "Non-ducted heating",
     "in.hvac_heating_type",
     "No"],
    [16, "Air Sealing + Attic + Duct Sealing + Drill and Fill Wall Insulation", "Envelope Package",
     "Ducted heating + Wood Frame walls",
     "Non-ducted OR non-wood frame",
     "in.hvac_heating_type, in.geometry_wall_type",
     "No"],
    [17, "EnergyStar Windows", "Envelope",
     "Buildings without triple-pane",
     "Already has Triple windows",
     "in.windows",
     "No"],
    [18, "Package: Dual-Speed GHP + Air Sealing + Attic + Duct Sealing", "HVAC+Envelope Package",
     "Ducted heating buildings",
     "Non-ducted heating",
     "in.hvac_heating_type",
     "Yes"],
    [19, "Electric Vehicle Adoption with Level 1 Charging", "EV",
     "All buildings", "(none)", "(none)", "No"],
    [20, "Electric Vehicle Adoption with Level 2 Charging", "EV",
     "All buildings", "(none)", "(none)", "No"],
    [21, "Efficient Electric Vehicle Adoption with Level 2 Charging", "EV",
     "All buildings", "(none)", "(none)", "No"],
    [22, "EV Adoption with Level 2 Charging and Demand Flexibility", "EV",
     "All buildings", "(none)", "(none)", "No"],
    [23, "Efficient EV Adoption with Level 2 Charging and Demand Flexibility", "EV",
     "All buildings", "(none)", "(none)", "No"],
    [24, "HVAC DF: On-peak Load Shedding, 2F Offset", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [25, "HVAC DF: Pre-peak Load Shifting, 2F Offset, 4hr Duration", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [26, "HVAC DF: On-peak Load Shedding, 4F Offset", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [27, "HVAC DF: Pre-peak Load Shifting, 4F Offset, 4hr Duration", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [28, "HVAC DF: Pre-peak Load Shifting, 2F Offset, 1hr Duration", "Demand Flex",
     "All buildings", "(none)", "(none)", "No"],
    [29, "General Air Sealing", "Envelope",
     "All buildings", "(none)", "(none)", "No"],
    [30, "General Air Sealing + Attic Floor Insulation + Duct Sealing", "Envelope",
     "Ducted heating buildings",
     "Non-ducted heating",
     "in.hvac_heating_type",
     "No"],
    [31, "Dual-Speed GHP + General Air Sealing + Attic + Duct Sealing", "HVAC+Envelope Package",
     "Ducted heating buildings",
     "Non-ducted heating",
     "in.hvac_heating_type",
     "Yes"],
    [32, "Variable-Speed GHP + General Air Sealing + Attic + Duct Sealing", "HVAC+Envelope Package",
     "Ducted heating buildings",
     "Non-ducted heating",
     "in.hvac_heating_type",
     "Yes"],
]

PACKAGE_HEADERS = [
    "Package ID",
    "Package Name",
    "Dataset",
    "Constituent Upgrade IDs",
    "Constituent Upgrade Names",
]

PACKAGE_DATA = [
    [54, "Package 1: Wall + Roof Insulation + New Windows", "ComStock", "48, 49, 52",
     "Wall Insulation, Roof Insulation, New Windows"],
    [55, "Package 2: LED + VS HP RTU or HP Boilers", "ComStock", "43, 1, 15",
     "LED Lighting, Variable Speed HP RTU, HP Boiler"],
    [56, "Package 3: Package 2 with Std HP RTU", "ComStock", "43, 4",
     "LED Lighting, Standard Performance HP RTU"],
    [57, "Package 4: Package 1 + Package 2", "ComStock", "48, 49, 52, 43, 1, 15",
     "Wall Insulation, Roof Insulation, New Windows, LED Lighting, VS HP RTU, HP Boiler"],
    [58, "Package 5: VS HP RTU/HP Boilers + Controls", "ComStock", "1, 15, 19, 20, 21",
     "VS HP RTU, HP Boiler, Air Side Economizers, DCV, Energy Recovery"],
    [59, "Package 6: GHP Variants", "ComStock", "28, 29, 30",
     "Hydronic W-to-W GHP, Packaged W-to-A GHP, Console W-to-A GHP"],
    [60, "Package 7: DF Lighting + Thermostat, Grid Peak + PV", "ComStock",
     "36, 37, 46", "DF Thermostat + Lighting Load Shed (Grid), PV"],
    [61, "Package 8: DF Lighting + Thermostat, Daily Bldg Peak + PV", "ComStock",
     "32, 34, 46", "DF Thermostat + Lighting Load Shed (Daily Bldg), PV"],
    [62, "Package 9: DF Lighting + Thermostat, Daily Bldg Peak", "ComStock",
     "32, 34", "DF Thermostat + Lighting Load Shed (Daily Bldg)"],
    [63, "Package 10: Package 1 + Package 6", "ComStock", "48, 49, 52, 28",
     "Wall Insulation, Roof Insulation, New Windows, Hydronic W-to-W GHP"],
    [64, "Package 11: Package 10 + LED Lighting", "ComStock", "48, 49, 52, 28, 43",
     "Wall Insulation, Roof Insulation, New Windows, Hydronic W-to-W GHP, LED Lighting"],
    [65, "Package 12: PV 40pct Roof + Roof Insulation", "ComStock", "46, 49",
     "Photovoltaics with 40pct Roof Coverage, Roof Insulation"],
    [18, "Package: Dual-Speed GHP + Air Sealing + Attic + Duct Sealing", "ResStock",
     "7, 11, 12, 13", "Dual-Speed GHP, Air Sealing, Attic Insulation, Duct Sealing"],
    [31, "Dual-Speed GHP + General Air Sealing + Attic + Duct Sealing", "ResStock",
     "7, 29, 12, 30", "Dual-Speed GHP, General Air Sealing, Attic Insulation, Duct Sealing"],
    [32, "Variable-Speed GHP + General Air Sealing + Attic + Duct Sealing", "ResStock",
     "8, 29, 12, 30", "Variable-Speed GHP, General Air Sealing, Attic Insulation, Duct Sealing"],
]

# ComStock package upgrade IDs (for highlight in Sheet 1)
COMSTOCK_PACKAGE_IDS = set(range(54, 66))


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def style_header_row(ws, num_cols):
    """Apply header styling: bold, blue fill, border, freeze, filter."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def auto_fit_columns(ws, cap=60):
    """Approximate auto-fit: measure max content length per column, capped."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max_length + 4, cap)
        ws.column_dimensions[col_letter].width = adjusted


def write_data_sheet(ws, headers, data, package_ids=None):
    """Write headers + data rows with alternating colours and optional
    package highlighting."""
    # Headers
    ws.append(headers)

    # Data rows
    for row in data:
        ws.append(row)

    # Style data rows (row index 2 onwards)
    for row_idx in range(2, ws.max_row + 1):
        upgrade_id = ws.cell(row=row_idx, column=1).value
        is_package = package_ids and upgrade_id in package_ids
        is_alt = (row_idx - 2) % 2 == 1  # alternating, 0-indexed

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            # Fill: package yellow > alternating grey > white
            if is_package:
                cell.fill = PACKAGE_FILL
            elif is_alt:
                cell.fill = ALT_ROW_FILL

            # Wrap text for the longer condition columns (4, 5, 6)
            if col_idx in (4, 5, 6):
                cell.alignment = WRAP_ALIGNMENT
            else:
                cell.alignment = TOP_ALIGNMENT

    style_header_row(ws, len(headers))
    auto_fit_columns(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()

    # Sheet 1: ComStock
    ws_com = wb.active
    ws_com.title = "ComStock Applicability Rules"
    write_data_sheet(ws_com, COMSTOCK_HEADERS, COMSTOCK_DATA,
                     package_ids=COMSTOCK_PACKAGE_IDS)

    # Sheet 2: ResStock
    ws_res = wb.create_sheet("ResStock Applicability Rules")
    write_data_sheet(ws_res, RESSTOCK_HEADERS, RESSTOCK_DATA)

    # Sheet 3: Package Constituents
    ws_pkg = wb.create_sheet("Package Constituents")
    write_data_sheet(ws_pkg, PACKAGE_HEADERS, PACKAGE_DATA)

    # Save
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "Applicability_Rules.xlsx")
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")

    # Quick verification
    from openpyxl import load_workbook
    wb_check = load_workbook(output_path)
    print(f"Sheets: {wb_check.sheetnames}")
    for name in wb_check.sheetnames:
        ws = wb_check[name]
        print(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
