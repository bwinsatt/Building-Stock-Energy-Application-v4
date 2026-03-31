"""
Generate ComStock_ResStock_Model_Features.xlsx workbook with 7 sheets
documenting all model features, raw columns, exclusion rationale,
field mappings, and imputation defaults.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_sheet(ws, wrap_cols=None):
    """Apply header styling, alternating rows, freeze pane, filters, and auto-width."""
    if wrap_cols is None:
        wrap_cols = set()

    # Header row
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.border = THIN_BORDER
            col_letter = get_column_letter(cell.column)
            if cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

    # Auto-fit widths (approximate)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        adjusted = min(max_len + 4, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 12)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Sheet 1: ComStock Selected Features
# ---------------------------------------------------------------------------

COMSTOCK_FEATURES = [
    ("in.comstock_building_type_group", "building_type", "Commercial building type category", "Categorical",
     "Education, Food Sales, Food Service, Healthcare, Lodging, Mercantile/Retail, Mixed Use, Office, Other, Public Assembly, Religious Worship, Warehouse and Storage, Unassigned",
     "User-Provided"),
    ("in.sqft..ft2", "sqft", "Conditioned floor area", "Numeric", "> 0 ft\u00b2", "User-Provided"),
    ("in.number_stories", "num_stories", "Number of above-ground stories", "Categorical (string)",
     '"1" through "14", "15_25", "over_25"', "User-Provided (mapped to categorical bins)"),
    ("in.as_simulated_ashrae_iecc_climate_zone_2006", "(derived from zipcode)",
     "ASHRAE/IECC 2006 climate zone", "Categorical",
     "1A, 2A, 2B, 3A, 3B, 3C, 4A, 4B, 4C, 5A, 5B, 5C, 6A, 6B, 7, 8", "Derived from zipcode"),
    ("cluster_name", "(derived from zipcode)", "Geographic regional cluster", "Categorical",
     '~6-7 clusters (e.g., "the Greater New York City Area")', "Derived from zipcode"),
    ("in.vintage", "year_built", "Construction era category", "Categorical",
     "Before1946, 1946-1959, 1960-1969, 1970-1979, 1980-1989, 1990-1999, 2000-2003, 2004-2006, 2007-2009, 2010-2012, 2013-2015, 2016-2019",
     "User-Provided (year mapped to bin)"),
    ("in.heating_fuel", "heating_fuel", "Primary space heating fuel", "Categorical",
     "Electricity, NaturalGas, FuelOil, Propane, DistrictHeating, None", "User-Provided or Imputed"),
    ("in.service_water_heating_fuel", "dhw_fuel", "Domestic hot water fuel", "Categorical",
     "Electricity, NaturalGas, FuelOil, Propane, DistrictHeating", "User-Provided or Imputed"),
    ("in.hvac_category", "(derived from hvac_system_type)", "HVAC system category", "Categorical",
     "Small Packaged Unit, Residential Style Central Systems, Multizone CAV/VAV, Zone-by-Zone",
     "Imputed (auto-derived from system type)"),
    ("in.hvac_cool_type", "(derived from hvac_system_type)", "Cooling equipment type", "Categorical",
     "DX, WCC, ACC, District, HP, None", "Imputed (auto-derived from system type)"),
    ("in.hvac_heat_type", "(derived from hvac_system_type)", "Heating equipment type", "Categorical",
     "Furnace, Boiler, HP, Electric Resistance, District, None", "Imputed (auto-derived from system type)"),
    ("in.hvac_system_type", "hvac_system_type", "Full HVAC system configuration", "Categorical",
     '29 variants (e.g., "PSZ-AC with gas coil", "PVAV with gas heat with electric reheat")',
     "User-Provided or Imputed"),
    ("in.hvac_vent_type", "(derived from hvac_system_type)", "Ventilation system type", "Categorical",
     "Central Single-zone RTU, Central Multizone CAV/VAV, DOAS+Zone terminal equipment, Zone terminal equipment",
     "Imputed (auto-derived from system type)"),
    ("in.wall_construction_type", "wall_construction", "Exterior wall construction material", "Categorical",
     "Mass, Metal Building, SteelFramed, WoodFramed", "User-Provided or Imputed"),
    ("in.window_type", "window_type", "Window glazing type", "Categorical",
     "Single Clear, Single Low-E, Double Clear, Double Low-E Non-metal, Double Low-E Metal, Triple Low-E",
     "User-Provided or Imputed"),
    ("in.window_to_wall_ratio_category", "window_to_wall_ratio", "Window-to-wall area ratio range", "Categorical",
     "0-10%, 10-20%, 20-30%, 30-40%, 40-50%, 50-60%, 60-70%, 70-80%, 80-90%, 90-100%",
     "User-Provided or Imputed"),
    ("in.airtightness..m3_per_m2_h", "(auto-imputed)", "Building envelope air leakage rate", "Numeric",
     "Default: 0.5 m\u00b3/m\u00b2\u00b7h", "Imputed (fixed default)"),
    ("in.interior_lighting_generation", "lighting_type", "Interior lighting technology generation", "Categorical",
     "gen1_t12_incandescent, gen2_t8_halogen, gen3_cfl, gen4_led, gen5_led", "User-Provided or Imputed"),
    ("in.weekday_operating_hours..hr", "operating_hours", "Weekday daily operating hours", "Categorical (string)",
     "5.75 to 18.75 in 0.25-hr increments", "User-Provided or Imputed (mapped from hrs/week)"),
    ("in.building_subtype", "(auto-imputed)", "Detailed building subtype", "Categorical",
     '"NA" (default)', "Imputed (fixed default)"),
    ("in.aspect_ratio", "(auto-imputed)", "Building length-to-width ratio", "Categorical (string)",
     '"2" (default)', "Imputed (fixed default)"),
    ("in.energy_code_followed_during_last_hvac_replacement", "(derived from state + year)",
     "Energy code for last HVAC replacement", "Categorical", "Various ASHRAE/IECC codes",
     "Derived from state + year_built"),
    ("in.energy_code_followed_during_last_roof_replacement", "(derived from state + year)",
     "Energy code for last roof replacement", "Categorical", "Various ASHRAE/IECC codes",
     "Derived from state + year_built"),
    ("in.energy_code_followed_during_last_walls_replacement", "(derived from state + year)",
     "Energy code for last walls replacement", "Categorical", "Various ASHRAE/IECC codes",
     "Derived from state + year_built"),
    ("in.energy_code_followed_during_last_svc_water_htg_replacement", "(derived from state + year)",
     "Energy code for last water heater replacement", "Categorical", "Various ASHRAE/IECC codes",
     "Derived from state + year_built"),
    ("hdd65f", "(derived from zipcode)", "Annual Heating Degree Days (base 65\u00b0F)", "Numeric",
     "Varies by location (e.g., 500-10000)", "Derived from zipcode"),
    ("cdd65f", "(derived from zipcode)", "Annual Cooling Degree Days (base 65\u00b0F)", "Numeric",
     "Varies by location (e.g., 100-5000)", "Derived from zipcode"),
    ("in.tstat_clg_sp_f..f", "thermostat_cooling_setpoint", "Cooling thermostat setpoint", "Numeric (\u00b0F)",
     "Typically 70-80\u00b0F; 999 = no data \u2192 NaN", "Optional (User or NaN)"),
    ("in.tstat_htg_sp_f..f", "thermostat_heating_setpoint", "Heating thermostat setpoint", "Numeric (\u00b0F)",
     "Typically 60-72\u00b0F; 999 = no data \u2192 NaN", "Optional (User or NaN)"),
    ("in.tstat_clg_delta_f..delta_f", "thermostat_cooling_setback",
     "Cooling setpoint offset during unoccupied", "Numeric (\u00b0F)",
     "Typically 0-10\u00b0F; 999 = no data \u2192 NaN", "Optional (User or NaN)"),
    ("in.tstat_htg_delta_f..delta_f", "thermostat_heating_setback",
     "Heating setpoint offset during unoccupied", "Numeric (\u00b0F)",
     "Typically 0-10\u00b0F; 999 = no data \u2192 NaN", "Optional (User or NaN)"),
    ("in.weekend_operating_hours..hr", "weekend_operating_hours", "Weekend daily operating hours",
     "Categorical (string)", "Similar range to weekday", "Optional (User or NaN)"),
    ("floor_plate_sqft", "(derived: sqft / stories)", "Floor area per story", "Numeric",
     "sqft / num_stories", "Derived"),
    ("baseline_electricity", "(from baseline model)", "Baseline electricity EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
    ("baseline_natural_gas", "(from baseline model)", "Baseline natural gas EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
    ("baseline_fuel_oil", "(from baseline model)", "Baseline fuel oil EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
    ("baseline_propane", "(from baseline model)", "Baseline propane EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
    ("baseline_district_heating", "(from baseline model)", "Baseline district heating EUI",
     "Numeric (kWh/ft\u00b2)", "\u2265 0", "Derived (model prediction)"),
]

# ---------------------------------------------------------------------------
# Sheet 2: ResStock Selected Features
# ---------------------------------------------------------------------------

RESSTOCK_FEATURES = [
    ("in.geometry_building_type_recs", "building_type", "RECS residential building type", "Categorical",
     "Multi-Family with 5+ Units", "User-Provided"),
    ("in.geometry_building_type_height", "(derived from stories)", "Building height category", "Categorical",
     "Low-Rise, Mid-Rise, High-Rise", "Derived from num_stories"),
    ("in.sqft..ft2", "sqft", "Conditioned floor area (per-unit for MF)", "Numeric",
     "300-6400 ft\u00b2 per unit", "User-Provided (converted to per-unit)"),
    ("in.geometry_stories", "num_stories", "Number of stories", "Categorical (string)",
     '"1" through "10+"', "User-Provided"),
    ("in.ashrae_iecc_climate_zone_2004", "(derived from zipcode)", "ASHRAE/IECC 2004 climate zone",
     "Categorical", "1A through 8", "Derived from zipcode"),
    ("cluster_name", "(derived from zipcode)", "Geographic regional cluster", "Categorical",
     "Regional clusters", "Derived from zipcode"),
    ("in.vintage", "year_built", "Construction era category", "Categorical",
     "<1940, 1940s, 1950s, 1960s, 1970s, 1980s, 1990s, 2000s, 2010s, 2020s",
     "User-Provided (year mapped to decade)"),
    ("in.heating_fuel", "heating_fuel", "Primary heating fuel", "Categorical",
     "Electricity, Natural Gas, Fuel Oil, Propane, Other Fuel, None", "User-Provided or Imputed"),
    ("in.hvac_heating_type", "hvac_system_type", "Heating system type", "Categorical",
     "Ducted Heating, Ducted Heat Pump, Non-Ducted Heating, Non-Ducted Heat Pump",
     "User-Provided (mapped)"),
    ("in.hvac_heating_efficiency", "(context-imputed)", "Heating system efficiency rating", "Categorical",
     'e.g., "Fuel Furnace, 80% AFUE", "ASHP, SEER 15, 8.5 HSPF"',
     "Imputed (based on fuel + system type)"),
    ("in.hvac_cooling_type", "(context-imputed)", "Cooling system type", "Categorical",
     "Central AC, Room AC, Ducted Heat Pump, Non-Ducted Heat Pump, None",
     "Imputed (based on system type)"),
    ("in.hvac_cooling_efficiency", "(context-imputed)", "Cooling system efficiency rating", "Categorical",
     'e.g., "AC, SEER 13", "Heat Pump, SEER 15"', "Imputed (based on system type)"),
    ("in.hvac_has_shared_system", "(context-imputed)", "Whether building has shared/central HVAC", "Categorical",
     "None, Fan Coil Units, Baseboards / Radiators", 'Imputed (default: "None")'),
    ("in.water_heater_fuel", "(context-imputed)", "Water heater fuel type", "Categorical",
     "Natural Gas, Electricity, Fuel Oil, Propane, Other Fuel", 'Imputed (default: "Natural Gas")'),
    ("in.water_heater_efficiency", "(context-imputed)", "Water heater efficiency class", "Categorical",
     'e.g., "Natural Gas Standard", "Electric Standard", "Heat Pump"', "Imputed (based on fuel)"),
    ("in.geometry_wall_type", "wall_construction", "Exterior wall material", "Categorical",
     "Wood Frame, Concrete Masonry Unit, Steel Frame", "User-Provided or Imputed"),
    ("in.insulation_wall", "(context-imputed)", "Wall insulation level", "Categorical",
     'e.g., "Wood Stud, R-7", "Wood Stud, R-13"', "Imputed (based on wall type)"),
    ("in.insulation_floor", "(auto-imputed)", "Floor insulation level", "Categorical",
     'Uninsulated, various R-values', 'Imputed (default: "Uninsulated")'),
    ("in.window_areas", "(auto-imputed)", "Window area distribution", "Categorical",
     'e.g., "F15 B15 L15 R15"', 'Imputed (default: "F15 B15 L15 R15")'),
    ("in.windows", "window_type", "Window type and performance", "Categorical",
     "Single-pane, Double-pane, Triple-pane variants", "User-Provided or Imputed"),
    ("in.lighting", "lighting_type", "Interior lighting type", "Categorical",
     "100% CFL, 100% LED, 100% Incandescent, various mixes", "User-Provided or Imputed"),
    ("in.infiltration", "(auto-imputed)", "Air leakage rate", "Categorical",
     'e.g., "5 ACH50", "10 ACH50", "3 ACH50"', 'Imputed (default: "5 ACH50")'),
    ("in.geometry_foundation_type", "(auto-imputed)", "Foundation type", "Categorical",
     "Slab, Heated Basement, Unheated Basement, Vented Crawlspace, etc.",
     'Imputed (default: "Slab")'),
    ("hdd65f", "(derived from zipcode)", "Annual Heating Degree Days (base 65\u00b0F)", "Numeric",
     "Varies by location", "Derived from zipcode"),
    ("cdd65f", "(derived from zipcode)", "Annual Cooling Degree Days (base 65\u00b0F)", "Numeric",
     "Varies by location", "Derived from zipcode"),
    ("in.heating_setpoint", "thermostat_heating_setpoint", "Heating setpoint", "Categorical (string)",
     'e.g., "60F", "68F", "72F"', "Optional (User or NaN)"),
    ("in.cooling_setpoint", "thermostat_cooling_setpoint", "Cooling setpoint", "Categorical (string)",
     'e.g., "75F", "78F", "80F"', "Optional (User or NaN)"),
    ("in.heating_setpoint_offset_magnitude", "thermostat_heating_setback",
     "Heating setpoint setback magnitude", "Categorical (string)",
     '"0F", "3F", "6F", "12F"', "Optional (User or NaN)"),
    ("in.cooling_setpoint_offset_magnitude", "thermostat_cooling_setback",
     "Cooling setpoint setback magnitude", "Categorical (string)",
     '"0F", "2F", "5F", "9F"', "Optional (User or NaN)"),
    ("floor_plate_sqft", "(derived: sqft / stories)", "Floor area per story", "Numeric",
     "sqft / num_stories", "Derived"),
    ("baseline_electricity", "(from baseline model)", "Baseline electricity EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
    ("baseline_natural_gas", "(from baseline model)", "Baseline natural gas EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
    ("baseline_fuel_oil", "(from baseline model)", "Baseline fuel oil EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
    ("baseline_propane", "(from baseline model)", "Baseline propane EUI", "Numeric (kWh/ft\u00b2)",
     "\u2265 0", "Derived (model prediction)"),
]


# ---------------------------------------------------------------------------
# Sheet 3: ComStock All Raw Columns
# ---------------------------------------------------------------------------

COMSTOCK_SELECTED = {
    "in.comstock_building_type_group", "in.sqft..ft2", "in.number_stories",
    "in.as_simulated_ashrae_iecc_climate_zone_2006", "in.vintage", "in.heating_fuel",
    "in.service_water_heating_fuel", "in.hvac_category", "in.hvac_cool_type",
    "in.hvac_heat_type", "in.hvac_system_type", "in.hvac_vent_type",
    "in.wall_construction_type", "in.window_type", "in.window_to_wall_ratio_category",
    "in.airtightness..m3_per_m2_h", "in.interior_lighting_generation",
    "in.weekday_operating_hours..hr", "in.building_subtype", "in.aspect_ratio",
    "in.energy_code_followed_during_last_hvac_replacement",
    "in.energy_code_followed_during_last_roof_replacement",
    "in.energy_code_followed_during_last_walls_replacement",
    "in.energy_code_followed_during_last_svc_water_htg_replacement",
    "in.tstat_clg_sp_f..f", "in.tstat_htg_sp_f..f",
    "in.tstat_clg_delta_f..delta_f", "in.tstat_htg_delta_f..delta_f",
    "in.weekend_operating_hours..hr",
}

COMSTOCK_EXCLUDED = {
    "in.as_simulated_census_division_name": "Geographic (aggregated to cluster_name)",
    "in.as_simulated_nhgis_county_gisjoin": "Identifier (used to derive cluster_name, then dropped)",
    "in.as_simulated_nhgis_puma_gisjoin": "Identifier (too granular geographic)",
    "in.as_simulated_nhgis_state_gisjoin": "Identifier (state captured via zipcode lookup)",
    "in.as_simulated_nhgis_tract_gisjoin": "Identifier (too granular geographic)",
    "in.as_simulated_state_name": "Redundant (state derived from zipcode)",
    "in.as_simulated_weather_file_2018": "Simulation artifact (weather file path)",
    "in.as_simulated_weather_file_tmy3": "Simulation artifact (weather file path)",
    "in.census_division_name": "Geographic (aggregated to cluster_name)",
    "in.census_region_name": "Geographic (aggregated to cluster_name)",
    "in.comstock_building_type": "Redundant (building_type_group is the grouped version used)",
    "in.economizer_changeover_temperature_fault_applicable": "Simulation artifact (fault modeling flag)",
    "in.economizer_damper_stuck_fault_applicable": "Simulation artifact (fault modeling flag)",
    "in.economizer_damper_stuck_fault_timing": "Simulation artifact (fault modeling detail)",
    "in.energy_code_followed_during_last_ext_lighting_replacement": "Unconfirmed",
    "in.energy_code_followed_during_last_int_equipment_replacement": "Unconfirmed",
    "in.energy_code_followed_during_original_building_construction": "Unconfirmed",
    "in.hvac_combined_type": "Redundant (hvac_system_type + hvac_category capture same info)",
    "in.hvac_night_variability": "Unconfirmed",
    "in.nhgis_state_gisjoin": "Identifier (state derived from zipcode)",
    "in.number_of_stories": "Redundant (in.number_stories is the selected version)",
    "in.ownership_type": "Unconfirmed",
    "in.party_responsible_for_operation": "Unconfirmed",
    "in.purchase_input_responsibility": "Unconfirmed",
    "in.rotation..degrees": "Simulation artifact (building orientation in simulation)",
    "in.sampling_region_id": "Identifier (sampling metadata)",
    "in.size_bin_id": "Identifier (sampling metadata)",
    "in.state": "Redundant (state derived from zipcode; climate zone + cluster used instead)",
    "in.state_name": "Redundant (same as in.state)",
    "in.tstat_at_least_one_clg_setback": "Redundant (setback delta captures this)",
    "in.tstat_at_least_one_htg_setback": "Redundant (setback delta captures this)",
    "in.tstat_has_clg_setback_all": "Redundant (setback delta captures this)",
    "in.tstat_has_htg_setback_all": "Redundant (setback delta captures this)",
    "in.upgrade_name": "Metadata (upgrade description, not a building feature)",
    "in.weekday_opening_time..hr": "Unconfirmed",
    "in.weekend_opening_time..hr": "Unconfirmed",
    "in.year_built": "Redundant (mapped to in.vintage categorical bins)",
}

COMSTOCK_ALL_RAW = sorted(list(COMSTOCK_SELECTED) + list(COMSTOCK_EXCLUDED.keys()))

# ---------------------------------------------------------------------------
# Sheet 4: ResStock All Raw Columns
# ---------------------------------------------------------------------------

RESSTOCK_SELECTED = {
    "in.geometry_building_type_recs", "in.geometry_building_type_height", "in.sqft..ft2",
    "in.geometry_stories", "in.ashrae_iecc_climate_zone_2004", "in.vintage",
    "in.heating_fuel", "in.hvac_heating_type", "in.hvac_heating_efficiency",
    "in.hvac_cooling_type", "in.hvac_cooling_efficiency", "in.hvac_has_shared_system",
    "in.water_heater_fuel", "in.water_heater_efficiency", "in.geometry_wall_type",
    "in.insulation_wall", "in.insulation_floor", "in.window_areas", "in.windows",
    "in.lighting", "in.infiltration", "in.geometry_foundation_type",
    "in.cooling_setpoint", "in.heating_setpoint",
    "in.cooling_setpoint_offset_magnitude", "in.heating_setpoint_offset_magnitude",
}

RESSTOCK_EXCLUDED = {
    # Geographic identifiers
    "in.ahs_region": "Geographic (aggregated to cluster_name)",
    "in.county": "Geographic (aggregated to cluster_name)",
    "in.county_and_puma": "Geographic (aggregated to cluster_name)",
    "in.county_metro_status": "Geographic (aggregated to cluster_name)",
    "in.county_name": "Geographic (aggregated to cluster_name)",
    "in.city": "Geographic (aggregated to cluster_name)",
    "in.census_division": "Geographic (aggregated to cluster_name)",
    "in.census_division_recs": "Geographic (aggregated to cluster_name)",
    "in.census_region": "Geographic (aggregated to cluster_name)",
    "in.custom_state": "Geographic (aggregated to cluster_name)",
    "in.location_region": "Geographic (aggregated to cluster_name)",
    "in.iso_rto_region": "Geographic (aggregated to cluster_name)",
    "in.metropolitan_and_micropolitan_statistical_area": "Geographic (aggregated to cluster_name)",
    "in.puma": "Geographic (aggregated to cluster_name)",
    "in.puma_metro_status": "Geographic (aggregated to cluster_name)",
    "in.reeds_balancing_area": "Geographic (aggregated to cluster_name)",
    "in.state": "Geographic (aggregated to cluster_name)",
    # Simulation artifacts
    "in.simulation_control_run_period_begin_day_of_month": "Simulation artifact",
    "in.simulation_control_run_period_begin_month": "Simulation artifact",
    "in.simulation_control_run_period_calendar_year": "Simulation artifact",
    "in.simulation_control_run_period_end_day_of_month": "Simulation artifact",
    "in.simulation_control_run_period_end_month": "Simulation artifact",
    "in.simulation_control_timestep": "Simulation artifact",
    "in.weather_file_city": "Simulation artifact",
    "in.weather_file_latitude": "Simulation artifact",
    "in.weather_file_longitude": "Simulation artifact",
    "in.hvac_cooling_autosizing_factor": "Simulation artifact",
    "in.hvac_heating_autosizing_factor": "Simulation artifact",
    "in.hvac_system_is_faulted": "Simulation artifact",
    "in.hvac_system_is_scaled": "Simulation artifact",
    "in.hvac_system_single_speed_ac_airflow": "Simulation artifact",
    "in.hvac_system_single_speed_ac_charge": "Simulation artifact",
    "in.hvac_system_single_speed_ashp_airflow": "Simulation artifact",
    "in.hvac_system_single_speed_ashp_charge": "Simulation artifact",
    # Socioeconomic
    "in.representative_income": "Socioeconomic (not relevant to energy physics)",
    "in.area_median_income": "Socioeconomic (not relevant to energy physics)",
    "in.federal_poverty_level": "Socioeconomic (not relevant to energy physics)",
    "in.income": "Socioeconomic (not relevant to energy physics)",
    "in.income_recs_2015": "Socioeconomic (not relevant to energy physics)",
    "in.income_recs_2020": "Socioeconomic (not relevant to energy physics)",
    "in.state_metro_median_income": "Socioeconomic (not relevant to energy physics)",
    "in.tenure": "Socioeconomic (not relevant to energy physics)",
    "in.household_has_tribal_persons": "Socioeconomic (not relevant to energy physics)",
    "in.aiannh_area": "Socioeconomic (not relevant to energy physics)",
    "in.vacancy_status": "Socioeconomic (not relevant to energy physics)",
    # Appliance-level
    "in.clothes_dryer": "Appliance-level detail (too granular for building-level EUI model)",
    "in.clothes_dryer_usage_level": "Appliance-level detail (too granular for building-level EUI model)",
    "in.clothes_washer": "Appliance-level detail (too granular for building-level EUI model)",
    "in.clothes_washer_presence": "Appliance-level detail (too granular for building-level EUI model)",
    "in.clothes_washer_usage_level": "Appliance-level detail (too granular for building-level EUI model)",
    "in.cooking_range": "Appliance-level detail (too granular for building-level EUI model)",
    "in.cooking_range_usage_level": "Appliance-level detail (too granular for building-level EUI model)",
    "in.dishwasher": "Appliance-level detail (too granular for building-level EUI model)",
    "in.dishwasher_usage_level": "Appliance-level detail (too granular for building-level EUI model)",
    "in.refrigerator": "Appliance-level detail (too granular for building-level EUI model)",
    "in.refrigerator_usage_level": "Appliance-level detail (too granular for building-level EUI model)",
    "in.ceiling_fan": "Appliance-level detail (too granular for building-level EUI model)",
    "in.dehumidifier": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_extra_refrigerator": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_freezer": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_gas_fireplace": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_gas_grill": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_gas_lighting": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_hot_tub_spa": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_pool": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_pool_heater": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_pool_pump": "Appliance-level detail (too granular for building-level EUI model)",
    "in.misc_well_pump": "Appliance-level detail (too granular for building-level EUI model)",
    "in.plug_loads": "Appliance-level detail (too granular for building-level EUI model)",
    "in.plug_load_diversity": "Appliance-level detail (too granular for building-level EUI model)",
    # Redundant
    "in.ashrae_iecc_climate_zone_2004_sub_cz_split": "Redundant (information captured by selected feature)",
    "in.building_america_climate_zone": "Redundant (information captured by selected feature)",
    "in.cec_climate_zone": "Redundant (information captured by selected feature)",
    "in.energystar_climate_zone_2023": "Redundant (information captured by selected feature)",
    "in.generation_and_emissions_assessment_region": "Redundant (information captured by selected feature)",
    "in.geometry_building_type_acs": "Redundant (information captured by selected feature)",
    "in.geometry_floor_area": "Redundant (information captured by selected feature)",
    "in.geometry_floor_area_bin": "Redundant (information captured by selected feature)",
    "in.geometry_stories_low_rise": "Redundant (information captured by selected feature)",
    "in.geometry_story_bin": "Redundant (information captured by selected feature)",
    "in.geometry_wall_exterior_finish": "Redundant (information captured by selected feature)",
    "in.hvac_heating_type_and_fuel": "Redundant (information captured by selected feature)",
    "in.vintage_acs": "Redundant (information captured by selected feature)",
    "in.upgrade_name": "Redundant (information captured by selected feature)",
    # EV-related
    "in.electric_vehicle_battery": "Unconfirmed",
    "in.electric_vehicle_charge_at_home": "Unconfirmed",
    "in.electric_vehicle_charger": "Unconfirmed",
    "in.electric_vehicle_miles_traveled": "Unconfirmed",
    "in.electric_vehicle_outlet_access": "Unconfirmed",
    "in.electric_vehicle_ownership": "Unconfirmed",
    # Building detail
    "in.bedrooms": "Unconfirmed",
    "in.corridor": "Unconfirmed",
    "in.door_area": "Unconfirmed",
    "in.doors": "Unconfirmed",
    "in.eaves": "Unconfirmed",
    "in.geometry_attic_type": "Unconfirmed",
    "in.geometry_building_horizontal_location_mf": "Unconfirmed",
    "in.geometry_building_horizontal_location_sfa": "Unconfirmed",
    "in.geometry_building_level_mf": "Unconfirmed",
    "in.geometry_building_number_units_mf": "Unconfirmed",
    "in.geometry_building_number_units_sfa": "Unconfirmed",
    "in.geometry_garage": "Unconfirmed",
    "in.geometry_space_combination": "Unconfirmed",
    "in.ground_thermal_conductivity": "Unconfirmed",
    "in.has_pv": "Unconfirmed",
    "in.overhangs": "Unconfirmed",
    "in.neighbors": "Unconfirmed",
    "in.orientation": "Unconfirmed",
    "in.radiant_barrier": "Unconfirmed",
    "in.roof_material": "Unconfirmed",
    # Secondary HVAC/mechanical
    "in.hvac_has_ducts": "Unconfirmed",
    "in.hvac_has_zonal_electric_heating": "Unconfirmed",
    "in.hvac_secondary_heating_efficiency": "Unconfirmed",
    "in.hvac_secondary_heating_fuel": "Unconfirmed",
    "in.hvac_secondary_heating_partial_space_conditioning": "Unconfirmed",
    "in.hvac_secondary_heating_type": "Unconfirmed",
    "in.hvac_shared_efficiencies": "Unconfirmed",
    "in.hvac_cooling_partial_space_conditioning": "Unconfirmed",
    "in.mechanical_ventilation": "Unconfirmed",
    # Water/DHW
    "in.hot_water_distribution": "Unconfirmed",
    "in.hot_water_fixtures": "Unconfirmed",
    "in.water_heater_in_unit": "Unconfirmed",
    "in.water_heater_location": "Unconfirmed",
    # Insulation details
    "in.insulation_ceiling": "Unconfirmed",
    "in.insulation_foundation_wall": "Unconfirmed",
    "in.insulation_rim_joist": "Unconfirmed",
    "in.insulation_roof": "Unconfirmed",
    "in.insulation_slab": "Unconfirmed",
    "in.interior_shading": "Unconfirmed",
    # Setpoint details already captured
    "in.cooling_setpoint_has_offset": "Redundant (setpoint magnitude captures this)",
    "in.cooling_setpoint_offset_period": "Redundant (setpoint magnitude captures this)",
    "in.heating_setpoint_has_offset": "Redundant (setpoint magnitude captures this)",
    "in.heating_setpoint_offset_period": "Redundant (setpoint magnitude captures this)",
    "in.cooling_unavailable_days": "Redundant (setpoint magnitude captures this)",
    "in.cooling_unavailable_period": "Redundant (setpoint magnitude captures this)",
    "in.heating_unavailable_days": "Redundant (setpoint magnitude captures this)",
    "in.heating_unavailable_period": "Redundant (setpoint magnitude captures this)",
    # Ventilation/usage
    "in.bathroom_spot_vent_hour": "Unconfirmed",
    "in.range_spot_vent_hour": "Unconfirmed",
    "in.natural_ventilation": "Unconfirmed",
    "in.usage_level": "Unconfirmed",
    "in.lighting_interior_use": "Unconfirmed",
    "in.lighting_other_use": "Unconfirmed",
    "in.holiday_lighting": "Unconfirmed",
    # Utility rates
    "in.utility_bill_electricity_fixed_charges": "Utility billing (not relevant to energy consumption physics)",
    "in.utility_bill_electricity_marginal_rates": "Utility billing (not relevant to energy consumption physics)",
    "in.utility_bill_fuel_oil_fixed_charges": "Utility billing (not relevant to energy consumption physics)",
    "in.utility_bill_fuel_oil_marginal_rates": "Utility billing (not relevant to energy consumption physics)",
    "in.utility_bill_natural_gas_fixed_charges": "Utility billing (not relevant to energy consumption physics)",
    "in.utility_bill_natural_gas_marginal_rates": "Utility billing (not relevant to energy consumption physics)",
    "in.utility_bill_propane_fixed_charges": "Utility billing (not relevant to energy consumption physics)",
    "in.utility_bill_propane_marginal_rates": "Utility billing (not relevant to energy consumption physics)",
    # Other
    "in.battery": "Unconfirmed",
    "in.occupants": "Unconfirmed",
    "in.pv_orientation": "Unconfirmed",
    "in.pv_system_size": "Unconfirmed",
    "in.units_represented": "Unconfirmed",
    "in.air_leakage_to_outside_ach50": "Unconfirmed",
    "in.duct_leakage_and_insulation": "Unconfirmed",
    "in.duct_location": "Unconfirmed",
    "in.electric_panel_breaker_space_total_count": "Unconfirmed",
    "in.electric_panel_service_rating..a": "Unconfirmed",
    "in.electric_panel_service_rating_bin..a": "Unconfirmed",
}

RESSTOCK_ALL_RAW = sorted(list(RESSTOCK_SELECTED) + list(RESSTOCK_EXCLUDED.keys()))


# ---------------------------------------------------------------------------
# Sheet 5: Excluded Column Rationale
# ---------------------------------------------------------------------------

EXCLUSION_RATIONALE = [
    ("Identifier/Metadata", "~6", "~3",
     "Building IDs, upgrade names, sampling IDs -- not building characteristics"),
    ("Geographic (aggregated to cluster)", "~5", "~17",
     "County, city, PUMA, state, census codes -- aggregated to regional cluster_name to reduce sparsity and prevent overfitting"),
    ("Simulation Artifact", "~6", "~17",
     "Weather file paths, rotation angles, fault flags, autosizing factors, run period settings -- artifacts of the EnergyPlus simulation, not real building characteristics"),
    ("Redundant", "~10", "~14",
     "Duplicate encodings (e.g., in.number_of_stories vs in.number_stories), alternate climate zone systems, or information already captured by a selected feature"),
    ("Socioeconomic", "0", "~11",
     "Income, tenure, poverty level, tribal status -- relevant for equity analysis but not for energy physics modeling"),
    ("Appliance-Level Detail", "0", "~25",
     "Individual appliance types, efficiency, and usage levels -- too granular for building-level EUI prediction"),
    ("Utility Billing", "0", "~8",
     "Fixed charges and marginal rates -- cost inputs, not energy consumption drivers"),
    ("Unconfirmed", "~6", "~68",
     "Column exists in raw data but exclusion reason not confirmed from code -- may have been excluded for feature importance, data quality, or model complexity reasons"),
]


# ---------------------------------------------------------------------------
# Sheet 6: Field Mappings
# ---------------------------------------------------------------------------

FIELD_MAPPINGS = [
    # ComStock wall_construction
    ("wall_construction", "Mass", "in.wall_construction_type", "Mass", "ComStock", "Direct mapping"),
    ("wall_construction", "Metal", "in.wall_construction_type", "Metal Building", "ComStock", "Direct mapping"),
    ("wall_construction", "SteelFrame", "in.wall_construction_type", "SteelFramed", "ComStock", "Direct mapping"),
    ("wall_construction", "WoodFrame", "in.wall_construction_type", "WoodFramed", "ComStock", "Direct mapping"),
    # ComStock lighting_type
    ("lighting_type", "T12", "in.interior_lighting_generation", "gen1_t12_incandescent", "ComStock", "Direct mapping"),
    ("lighting_type", "T8", "in.interior_lighting_generation", "gen2_t8_halogen", "ComStock", "Direct mapping"),
    ("lighting_type", "CFL", "in.interior_lighting_generation", "gen3_cfl", "ComStock", "Direct mapping"),
    ("lighting_type", "LED", "in.interior_lighting_generation", "gen4_led", "ComStock", "Direct mapping"),
    # ComStock window_to_wall_ratio
    ("window_to_wall_ratio", "10-20%", "in.window_to_wall_ratio_category", "20-30%", "ComStock",
     "Example binning -- user percentage mapped to nearest category"),
    # ComStock num_stories
    ("num_stories", "1 (integer)", "in.number_stories", '"1"', "ComStock",
     "Integer mapped to categorical string; values > 14 mapped to 15_25 or over_25"),
    ("num_stories", "20 (integer)", "in.number_stories", '"15_25"', "ComStock",
     "Integer 15-25 mapped to categorical bin"),
    ("num_stories", "30 (integer)", "in.number_stories", '"over_25"', "ComStock",
     "Integer > 25 mapped to categorical bin"),
    # ComStock year_built
    ("year_built", "1985 (integer)", "in.vintage", '"1980-1989"', "ComStock",
     "Year mapped to decade bin"),
    ("year_built", "1940 (integer)", "in.vintage", '"Before1946"', "ComStock",
     "Years before 1946 grouped into single bin"),
    ("year_built", "2018 (integer)", "in.vintage", '"2016-2019"', "ComStock",
     "Year mapped to appropriate bin"),
    # ComStock operating_hours
    ("operating_hours", "50 hrs/week", "in.weekday_operating_hours..hr", "10.0",
     "ComStock", "Divide by 5 weekdays, snap to nearest 0.25, clamp 5.75-18.75"),
    ("operating_hours", "80 hrs/week", "in.weekday_operating_hours..hr", "16.0",
     "ComStock", "Divide by 5 weekdays, snap to nearest 0.25, clamp 5.75-18.75"),
    # ResStock building_type
    ("building_type", 'Multi-Family (5+ Units)', "in.geometry_building_type_recs",
     "Multi-Family with 5+ Units", "ResStock", "Parenthetical to 'with' format"),
    # ResStock hvac_system_type
    ("hvac_system_type", "forced_air_furnace", "in.hvac_heating_type", "Ducted Heating",
     "ResStock", "Furnace/boiler types mapped to Ducted/Non-Ducted Heating"),
    ("hvac_system_type", "ducted_heat_pump", "in.hvac_heating_type", "Ducted Heat Pump",
     "ResStock", "Heat pump types mapped to Ducted/Non-Ducted Heat Pump"),
    ("hvac_system_type", "mini_split_heat_pump", "in.hvac_heating_type", "Non-Ducted Heat Pump",
     "ResStock", "Mini-split mapped to Non-Ducted Heat Pump"),
    # ResStock sqft
    ("sqft", "25000 (total building)", "in.sqft..ft2", "~500-1000 (per-unit)",
     "ResStock", "Total sqft divided by estimated unit count, clamped 300-6400"),
    # ResStock thermostat
    ("thermostat_heating_setpoint", "70.0", "in.heating_setpoint", '"70F"',
     "ResStock", "Float converted to categorical string with F suffix"),
    ("thermostat_cooling_setpoint", "75.0", "in.cooling_setpoint", '"75F"',
     "ResStock", "Float converted to categorical string with F suffix"),
    # ResStock setback
    ("thermostat_heating_setback", "5.0", "in.heating_setpoint_offset_magnitude", '"6F"',
     "ResStock", "Snapped to nearest category: [0, 3, 6, 12]\u00b0F"),
    ("thermostat_cooling_setback", "3.0", "in.cooling_setpoint_offset_magnitude", '"2F"',
     "ResStock", "Snapped to nearest category: [0, 2, 5, 9]\u00b0F"),
]


# ---------------------------------------------------------------------------
# Sheet 7: Imputation Defaults
# ---------------------------------------------------------------------------

IMPUTATION_DEFAULTS_HEADER = [
    "Building Type", "Dataset", "heating_fuel", "dhw_fuel", "hvac_system_type",
    "wall_construction", "window_type", "window_to_wall_ratio", "lighting_type",
    "operating_hours", "year_built",
]

IMPUTATION_DEFAULTS = [
    ("Office", "ComStock", "NaturalGas", "NaturalGas", "central_chiller_boiler",
     "Mass", "Double Low-E Non-metal", "20-30%", "LED", "50 hrs/week", "1985"),
    ("Mercantile/Retail", "ComStock", "NaturalGas", "NaturalGas", "packaged_rooftop",
     "WoodFrame", "Double Clear", "20-30%", "LED", "65 hrs/week", "1990"),
    ("Education", "ComStock", "NaturalGas", "NaturalGas", "central_chiller_boiler",
     "Mass", "Double Clear", "20-30%", "LED", "50 hrs/week", "1975"),
    ("Healthcare", "ComStock", "NaturalGas", "NaturalGas", "central_chiller_boiler",
     "Mass", "Double Low-E Non-metal", "20-30%", "LED", "100 hrs/week", "1990"),
    ("Lodging", "ComStock", "NaturalGas", "NaturalGas", "packaged_terminal",
     "Mass", "Double Clear", "20-30%", "LED", "168 hrs/week", "1985"),
    ("Food Service", "ComStock", "NaturalGas", "NaturalGas", "packaged_rooftop",
     "WoodFrame", "Double Clear", "10-20%", "LED", "80 hrs/week", "1990"),
    ("Food Sales", "ComStock", "Electricity", "Electricity", "packaged_rooftop",
     "Mass", "Double Clear", "10-20%", "LED", "80 hrs/week", "1995"),
    ("Warehouse and Storage", "ComStock", "NaturalGas", "NaturalGas", "packaged_rooftop",
     "Metal", "Single Clear", "0-10%", "LED", "50 hrs/week", "1990"),
    ("Public Assembly", "ComStock", "NaturalGas", "NaturalGas", "central_chiller_boiler",
     "Mass", "Double Clear", "20-30%", "LED", "40 hrs/week", "1980"),
    ("Religious Worship", "ComStock", "NaturalGas", "NaturalGas", "packaged_rooftop",
     "WoodFrame", "Double Clear", "20-30%", "CFL", "20 hrs/week", "1975"),
    ("Mixed Use", "ComStock", "NaturalGas", "NaturalGas", "packaged_rooftop",
     "Mass", "Double Clear", "20-30%", "LED", "60 hrs/week", "1985"),
    ("Other", "ComStock", "NaturalGas", "NaturalGas", "packaged_rooftop",
     "Mass", "Double Clear", "20-30%", "LED", "50 hrs/week", "1985"),
    ("Unassigned", "ComStock", "NaturalGas", "NaturalGas", "packaged_rooftop",
     "Mass", "Double Clear", "20-30%", "LED", "50 hrs/week", "1985"),
    ("Multi-Family (5+ Units)", "ResStock", "NaturalGas", "n/a", "forced_air_furnace",
     "WoodFrame", "Double-pane", "n/a", "100% LED", "n/a", "1985"),
]


# ===========================================================================
# Build the workbook
# ===========================================================================

def build_workbook():
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1: ComStock Selected Features
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "ComStock Selected Features"
    headers1 = ["Feature Name", "User-Facing Field", "Description", "Data Type",
                "Valid Values/Ranges", "Source"]
    ws1.append(headers1)
    for row in COMSTOCK_FEATURES:
        ws1.append(list(row))
    # Wrap: Description (col 3), Valid Values (col 5)
    style_sheet(ws1, wrap_cols={3, 5})

    # -----------------------------------------------------------------------
    # Sheet 2: ResStock Selected Features
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("ResStock Selected Features")
    ws2.append(headers1)
    for row in RESSTOCK_FEATURES:
        ws2.append(list(row))
    style_sheet(ws2, wrap_cols={3, 5})

    # -----------------------------------------------------------------------
    # Sheet 3: ComStock All Raw Columns
    # -----------------------------------------------------------------------
    ws3 = wb.create_sheet("ComStock All Raw Columns")
    headers3 = ["Column Name", "Selected for Model", "Exclusion Reason"]
    ws3.append(headers3)
    for col_name in COMSTOCK_ALL_RAW:
        if col_name in COMSTOCK_SELECTED:
            ws3.append([col_name, "Yes", ""])
        else:
            ws3.append([col_name, "No", COMSTOCK_EXCLUDED.get(col_name, "")])
    style_sheet(ws3, wrap_cols={3})

    # -----------------------------------------------------------------------
    # Sheet 4: ResStock All Raw Columns
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("ResStock All Raw Columns")
    ws4.append(headers3)
    for col_name in RESSTOCK_ALL_RAW:
        if col_name in RESSTOCK_SELECTED:
            ws4.append([col_name, "Yes", ""])
        else:
            ws4.append([col_name, "No", RESSTOCK_EXCLUDED.get(col_name, "")])
    style_sheet(ws4, wrap_cols={3})

    # -----------------------------------------------------------------------
    # Sheet 5: Excluded Column Rationale
    # -----------------------------------------------------------------------
    ws5 = wb.create_sheet("Excluded Column Rationale")
    headers5 = ["Exclusion Category", "Count (ComStock)", "Count (ResStock)", "Explanation"]
    ws5.append(headers5)
    for row in EXCLUSION_RATIONALE:
        ws5.append(list(row))
    style_sheet(ws5, wrap_cols={4})

    # -----------------------------------------------------------------------
    # Sheet 6: Field Mappings
    # -----------------------------------------------------------------------
    ws6 = wb.create_sheet("Field Mappings")
    headers6 = ["User Input Field", "User Value", "Model Feature", "Model Value",
                "Dataset", "Notes"]
    ws6.append(headers6)
    for row in FIELD_MAPPINGS:
        ws6.append(list(row))
    style_sheet(ws6, wrap_cols={6})

    # -----------------------------------------------------------------------
    # Sheet 7: Imputation Defaults
    # -----------------------------------------------------------------------
    ws7 = wb.create_sheet("Imputation Defaults")
    ws7.append(IMPUTATION_DEFAULTS_HEADER)
    for row in IMPUTATION_DEFAULTS:
        ws7.append(list(row))
    style_sheet(ws7, wrap_cols=set())

    return wb


def main():
    wb = build_workbook()
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "ComStock_ResStock_Model_Features.xlsx")
    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")

    # Quick summary
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  {sheet_name}: {ws.max_row - 1} data rows x {ws.max_column} columns")


if __name__ == "__main__":
    main()
