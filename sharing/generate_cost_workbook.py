"""
Generate an Excel workbook from upgrade_cost_lookup.json.

Creates sharing/Upgrade_Cost_Data.xlsx with three sheets:
  1. ComStock Upgrades
  2. ResStock Upgrades
  3. Metadata & Sources
"""

import json
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent
JSON_PATH = PROJECT_ROOT / "backend" / "app" / "data" / "upgrade_cost_lookup.json"
OUTPUT_PATH = Path(__file__).resolve().parent / "Upgrade_Cost_Data.xlsx"

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
HEADER_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALT_ROW_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

CONFIDENCE_FILLS = {
    "high": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    "medium": PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),
    "low": PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
}

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")

COST_FORMAT = "#,##0.00"

# Column definitions for upgrade sheets
COLUMNS = [
    "Upgrade ID",
    "Upgrade Name",
    "Measure Category",
    "Cost Low",
    "Cost Mid",
    "Cost High",
    "Cost Basis",
    "Cost Year ($)",
    "Useful Life (yrs)",
    "Cost Type",
    "Confidence",
    "Scout ECM Name",
    "Description",
    "Notes",
    "Sources",
]

# Columns that should use number format with 2 decimals
COST_COLUMNS = {"Cost Low", "Cost Mid", "Cost High"}

# Columns that should wrap text
WRAP_COLUMNS = {"Description", "Notes", "Sources"}

# Index of Confidence column (1-based)
CONFIDENCE_COL_NAME = "Confidence"


def format_sources(sources: list[dict]) -> str:
    """Combine source objects into a readable string."""
    parts = []
    for src in sources:
        title = src.get("title", "").strip()
        year = src.get("year")
        if title:
            entry = title
            if year:
                entry += f" ({year})"
            parts.append(entry)
    return "; ".join(parts) if parts else ""


def build_row(record: dict) -> list:
    """Build a data row from an upgrade record in column order."""
    return [
        record.get("upgrade_id"),
        record.get("upgrade_name", ""),
        record.get("measure_category", ""),
        record.get("cost_low"),
        record.get("cost_mid"),
        record.get("cost_high"),
        record.get("cost_basis", ""),
        record.get("cost_year_dollars"),
        record.get("useful_life_years"),
        record.get("cost_type", ""),
        record.get("confidence", ""),
        record.get("scout_ecm_name", ""),
        record.get("description", ""),
        record.get("notes", ""),
        format_sources(record.get("sources", [])),
    ]


def apply_header_style(ws, num_cols: int) -> None:
    """Style the header row (row 1)."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top")


def style_data_rows(ws, num_cols: int, start_row: int = 2) -> None:
    """Apply alternating row colors, cost formatting, wrap text, and
    confidence coloring to all data rows."""
    confidence_col_idx = COLUMNS.index(CONFIDENCE_COL_NAME) + 1  # 1-based

    for row_idx in range(start_row, ws.max_row + 1):
        is_alt = (row_idx - start_row) % 2 == 1
        base_fill = ALT_ROW_FILL if is_alt else WHITE_FILL

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = COLUMNS[col_idx - 1]

            # Number format for cost columns
            if col_name in COST_COLUMNS:
                cell.number_format = COST_FORMAT

            # Wrap text for long-text columns
            if col_name in WRAP_COLUMNS:
                cell.alignment = WRAP_ALIGNMENT
            else:
                cell.alignment = TOP_ALIGNMENT

            # Confidence conditional fill
            if col_idx == confidence_col_idx:
                val = str(cell.value).lower() if cell.value else ""
                cell.fill = CONFIDENCE_FILLS.get(val, base_fill)
            else:
                cell.fill = base_fill

            cell.border = THIN_BORDER


def auto_fit_columns(ws, max_width: int = 60) -> None:
    """Set column widths based on content, capped at max_width."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                # Approximate: take the longest line in the cell
                lines = str(cell.value).split("\n")
                longest = max(len(line) for line in lines)
                max_len = max(max_len, longest)
        # Add a small buffer
        adjusted = min(max_len + 3, max_width)
        # Minimum width of 10
        adjusted = max(adjusted, 10)
        ws.column_dimensions[col_letter].width = adjusted


def write_upgrade_sheet(ws, records: list[dict]) -> None:
    """Populate an upgrade sheet with headers, data, filters, freeze pane."""
    # Header
    ws.append(COLUMNS)
    num_cols = len(COLUMNS)

    # Sort by upgrade_id and write rows
    sorted_records = sorted(records, key=lambda r: r.get("upgrade_id", 0))
    for rec in sorted_records:
        ws.append(build_row(rec))

    # Styling
    apply_header_style(ws, num_cols)
    style_data_rows(ws, num_cols)
    auto_fit_columns(ws)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter on header row
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def write_metadata_sheet(ws, metadata: dict, all_sources: list[dict]) -> None:
    """Build the Metadata & Sources reference sheet."""
    bold = Font(bold=True, size=12)
    normal = Font(size=11)

    # Row 1-2: Cost Year Dollars
    ws.cell(row=1, column=1, value="Cost Year Dollars").font = bold
    ws.cell(row=1, column=2, value=metadata.get("cost_year_dollars", ""))

    # Row 3-4: Generated Date
    ws.cell(row=3, column=1, value="Generated Date").font = bold
    ws.cell(row=3, column=2, value=metadata.get("generated", ""))

    # Row 5+: Notes
    ws.cell(row=5, column=1, value="Notes").font = bold
    ws.cell(row=5, column=2, value=metadata.get("notes", ""))
    ws.cell(row=5, column=2).alignment = WRAP_ALIGNMENT

    # Sources summary
    sources_summary = metadata.get("sources_summary", {})
    row = 7
    ws.cell(row=row, column=1, value="Sources Summary").font = bold
    row += 1
    for key, val in sources_summary.items():
        ws.cell(row=row, column=1, value=key.replace("_", " ").title()).font = Font(bold=True)
        if isinstance(val, list):
            ws.cell(row=row, column=2, value="; ".join(val))
        else:
            ws.cell(row=row, column=2, value=str(val))
        row += 1

    # Blank row
    row += 1

    # Source Inventory header
    ws.cell(row=row, column=1, value="Source Inventory").font = Font(bold=True, size=13)
    row += 1

    # Sub-headers
    source_headers = ["Source Title", "Author", "Year", "URL"]
    for ci, header in enumerate(source_headers, start=1):
        cell = ws.cell(row=row, column=ci, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    row += 1

    # Deduplicate sources by title (primary key)
    seen_titles = set()
    unique_sources = []
    for src in all_sources:
        title = src.get("title", "").strip()
        if title and title not in seen_titles:
            seen_titles.add(title)
            unique_sources.append(src)

    # Sort by title
    unique_sources.sort(key=lambda s: s.get("title", "").lower())

    for src in unique_sources:
        ws.cell(row=row, column=1, value=src.get("title", ""))
        ws.cell(row=row, column=2, value=src.get("author", ""))
        ws.cell(row=row, column=3, value=src.get("year"))
        ws.cell(row=row, column=4, value=src.get("url", ""))
        # Alternating rows for source inventory
        is_alt = (row % 2) == 0
        fill = ALT_ROW_FILL if is_alt else WHITE_FILL
        for ci in range(1, 5):
            ws.cell(row=row, column=ci).fill = fill
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    # Auto-fit
    auto_fit_columns(ws)


def write_regional_adjustments_sheet(ws, data: dict) -> None:
    """Build the Regional Cost Adjustments sheet from the JSON data."""
    regional = data.get("metadata", {}).get("regional_adjustments", {})

    # Note row
    ws.cell(row=1, column=1, value=(
        "Regional cost adjustment multipliers by state, sourced from RSMeans "
        "city cost indices via NREL Scout. A multiplier of 1.0 = national average "
        "(Michigan baseline). Applied to all cost estimates (low/mid/high) at "
        "prediction time: Adjusted Cost = National Cost × Regional Factor."
    ))
    ws.cell(row=1, column=1).alignment = WRAP_ALIGNMENT
    ws.cell(row=1, column=1).font = Font(italic=True)
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 45

    # Headers at row 3
    headers = [
        "State", "Representative City", "Residential Factor",
        "Commercial Factor", "Residential vs National", "Commercial vs National",
    ]
    header_row = 3
    for ci, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=ci, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = f"A{header_row + 1}"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"

    # State data — from the JSON regional_adjustments or hardcoded from Scout
    # The JSON stores {state_abbr: {city, residential, commercial}}
    states = []
    if regional:
        for state_abbr, info in regional.items():
            states.append({
                "state": state_abbr,
                "city": info.get("city", ""),
                "residential": info.get("residential", 1.0),
                "commercial": info.get("commercial", 1.0),
            })
    else:
        # Fallback: hardcoded from Scout loc_cost_adj.csv
        _raw = [
            ("AL", "Huntsville", 0.83, 0.85), ("AK", "Anchorage", 1.17, 1.16),
            ("AZ", "Phoenix", 0.84, 0.87), ("AR", "Little Rock", 0.83, 0.82),
            ("CA", "Los Angeles", 1.15, 1.12), ("CO", "Denver", 0.91, 0.92),
            ("CT", "Bridgeport", 1.10, 1.08), ("DE", "Wilmington", 1.01, 1.04),
            ("DC", "Washington, D.C.", 0.92, 0.96), ("FL", "Jacksonville", 0.81, 0.84),
            ("GA", "Atlanta", 0.90, 0.89), ("HI", "Honolulu", 1.22, 1.19),
            ("ID", "Boise", 0.89, 0.91), ("IL", "Chicago", 1.25, 1.20),
            ("IN", "Indianapolis", 0.92, 0.92), ("IA", "Des Moines", 0.92, 0.94),
            ("KS", "Wichita", 0.81, 0.86), ("KY", "Louisville", 0.89, 0.88),
            ("LA", "New Orleans", 0.85, 0.85), ("ME", "Portland", 0.94, 0.94),
            ("MD", "Baltimore", 0.93, 0.94), ("MA", "Boston", 1.18, 1.14),
            ("MI", "Detroit", 1.00, 1.00), ("MN", "Minneapolis", 1.09, 1.07),
            ("MS", "Jackson", 0.84, 0.85), ("MO", "Kansas City", 0.99, 0.99),
            ("MT", "Billings", 0.89, 0.91), ("NE", "Omaha", 0.90, 0.92),
            ("NV", "Las Vegas", 1.03, 1.05), ("NH", "Manchester", 0.99, 0.97),
            ("NJ", "Newark", 1.20, 1.17), ("NM", "Albuquerque", 0.86, 0.87),
            ("NY", "New York", 1.36, 1.32), ("NC", "Charlotte", 0.99, 0.87),
            ("ND", "Fargo", 0.87, 0.91), ("OH", "Columbus", 0.91, 0.92),
            ("OK", "Oklahoma City", 0.84, 0.85), ("OR", "Portland", 1.02, 1.03),
            ("PA", "Philadelphia", 1.17, 1.16), ("RI", "Providence", 1.09, 1.06),
            ("SC", "Charleston", 0.98, 0.85), ("SD", "Sioux Falls", 0.92, 0.92),
            ("TN", "Nashville", 0.87, 0.89), ("TX", "Houston", 0.84, 0.87),
            ("UT", "Salt Lake City", 0.85, 0.91), ("VT", "Burlington", 1.00, 0.95),
            ("VA", "Newport News", 0.97, 0.87), ("WA", "Seattle", 1.06, 1.07),
            ("WV", "Charleston", 0.94, 0.94), ("WI", "Milwaukee", 1.07, 1.04),
            ("WY", "Cheyenne", 0.86, 0.89),
        ]
        for s, c, r, cm in _raw:
            states.append({"state": s, "city": c, "residential": r, "commercial": cm})

    states.sort(key=lambda x: x["state"])

    # Write data rows
    row = header_row + 1
    for entry in states:
        res_factor = entry["residential"]
        com_factor = entry["commercial"]
        ws.cell(row=row, column=1, value=entry["state"])
        ws.cell(row=row, column=2, value=entry["city"])
        c_res = ws.cell(row=row, column=3, value=res_factor)
        c_com = ws.cell(row=row, column=4, value=com_factor)
        c_res.number_format = "0.00"
        c_com.number_format = "0.00"

        # % vs national (1.0 baseline)
        res_pct = (res_factor - 1.0) * 100
        com_pct = (com_factor - 1.0) * 100
        c_res_pct = ws.cell(row=row, column=5, value=res_pct / 100)
        c_com_pct = ws.cell(row=row, column=6, value=com_pct / 100)
        c_res_pct.number_format = "+0.0%;-0.0%;0.0%"
        c_com_pct.number_format = "+0.0%;-0.0%;0.0%"

        # Alternating fill + borders
        is_alt = (row - header_row - 1) % 2 == 1
        fill = ALT_ROW_FILL if is_alt else WHITE_FILL
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=ci)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = TOP_ALIGNMENT

        row += 1

    auto_fit_columns(ws)


def main() -> None:
    # Load data
    with open(JSON_PATH, "r") as f:
        data = json.load(f)

    metadata = data.get("metadata", {})
    comstock_records = list(data.get("comstock", {}).values())
    resstock_records = list(data.get("resstock", {}).values())

    # Collect all sources across both datasets
    all_sources = []
    for rec in comstock_records + resstock_records:
        all_sources.extend(rec.get("sources", []))

    # Create workbook
    wb = Workbook()

    # Sheet 1: ComStock Upgrades
    ws_com = wb.active
    ws_com.title = "ComStock Upgrades"
    write_upgrade_sheet(ws_com, comstock_records)

    # Sheet 2: ResStock Upgrades
    ws_res = wb.create_sheet("ResStock Upgrades")
    write_upgrade_sheet(ws_res, resstock_records)

    # Sheet 3: Regional Cost Adjustments
    ws_reg = wb.create_sheet("Regional Cost Adjustments")
    write_regional_adjustments_sheet(ws_reg, data)

    # Sheet 4: Metadata & Sources
    ws_meta = wb.create_sheet("Metadata & Sources")
    write_metadata_sheet(ws_meta, metadata, all_sources)

    # Save
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"Workbook saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
