"""Tests for the carbon performance export service (openpyxl cell writing)."""

import io
import re
import zipfile

import openpyxl

from app.schemas.request import BuildingInput
from app.schemas.response import (
    BuildingResult, BaselineResult, FuelBreakdown, MeasureResult,
    CostEstimate, InputSummary,
)
from app.services import export_service


def assert_valid_xlsx(data: bytes):
    """Catch the corruption class Excel flags but openpyxl tolerates.

    Verifies the OOXML package has no external links and no dangling
    relationship references (every r:id in a part exists in its .rels).
    """
    z = zipfile.ZipFile(io.BytesIO(data))
    names = z.namelist()

    assert not any("externalLink" in n for n in names), "external links present"
    assert "externalReference" not in z.read("xl/workbook.xml").decode()

    def rels_for(part):
        head, _, tail = part.rpartition("/")
        relp = f"{head}/_rels/{tail}.rels" if head else f"_rels/{tail}.rels"
        try:
            return z.read(relp).decode()
        except KeyError:
            return ""

    dangling = []
    for n in names:
        if n.endswith(".xml") and not n.endswith(".rels"):
            body = z.read(n).decode(errors="ignore")
            ids = set(re.findall(r'r:id="([^"]+)"', body))
            defined = set(re.findall(r'Id="([^"]+)"', rels_for(n)))
            if ids - defined:
                dangling.append((n, ids - defined))
    assert not dangling, f"dangling relationship refs: {dangling}"

    # Must still reopen cleanly
    openpyxl.load_workbook(io.BytesIO(data))


def _fake_result():
    baseline = BaselineResult(
        total_eui_kbtu_sf=100.0,
        eui_by_fuel=FuelBreakdown(
            electricity=34.12, natural_gas=50.0, fuel_oil=0.0, propane=0.0,
        ),
    )
    measure = MeasureResult(
        upgrade_id=7, name="LED Lighting", category="lighting", applicable=True,
        cost=CostEstimate(
            installed_cost_per_sf=2.0, installed_cost_total=100000.0,
            cost_range={"low": 90000.0, "high": 110000.0}, useful_life_years=15,
            regional_factor=1.0, confidence="high",
        ),
        electricity_savings_kwh=1.0,      # kWh/sf
        gas_savings_therms=0.1,           # therms/sf
        savings_by_fuel=FuelBreakdown(
            electricity=3.412, natural_gas=2.931, fuel_oil=0.0, propane=0.0,
        ),
    )
    summary = InputSummary(
        climate_zone="4A", cluster_name="x", state="NY",
        vintage_bucket="<1946", imputed_fields=[],
    )
    result = BuildingResult(
        building_index=0, baseline=baseline, measures=[measure], input_summary=summary
    )
    rates = {"electricity": 0.20, "natural_gas": 0.05}
    return result, rates


def _building():
    return BuildingInput(
        building_type="Office", sqft=1000, num_stories=3,
        zipcode="00501", year_built=1985,
    )


def test_build_workbook_fills_expected_cells(monkeypatch):
    monkeypatch.setattr(export_service, "assess_building_for_export", lambda *a, **k: _fake_result())
    monkeypatch.setattr(export_service, "get_egrid_subregion", lambda z: "NYLI")

    data = export_service.build_carbon_performance_workbook(
        _building(), [7], model_manager=None, cost_calculator=None, espm_property_type="Office",
    )
    ws = openpyxl.load_workbook(io.BytesIO(data))["Export BlueLynx"]

    # Measure row 6
    assert ws["C6"].value == "LED Lighting"
    assert ws["D6"].value == 100000.0
    assert ws["L6"].value == 1000           # 1.0 kWh/sf * 1000 sf
    assert ws["H6"].value == 200.0          # 1000 kWh * $0.20
    assert ws["M6"].value == 100.0          # 0.1 therms/sf * 1000 sf
    assert round(ws["I6"].value, 0) == 147  # 100 therms * 29.31 kWh/therm * $0.05

    # Baseline metrics
    assert ws["AJ8"].value == "00501"
    assert ws["AJ9"].value == "Office"
    assert ws["AJ11"].value == "NY"
    assert ws["AJ14"].value == "NYLI"
    assert ws["AJ15"].value == "4A"
    assert ws["AJ16"].value == 10000        # 34.12/3.412*1000
    assert ws["AJ17"].value == 2000.0       # 10000 kWh * $0.20
    assert ws["AJ20"].value == 500.0        # 50.0/100*1000
    assert ws["AJ24"].value == 1000
    assert ws["AJ32"].value == "NREL Cambium, 2022"


def test_generated_workbook_is_not_corrupt(monkeypatch):
    monkeypatch.setattr(export_service, "assess_building_for_export", lambda *a, **k: _fake_result())
    monkeypatch.setattr(export_service, "get_egrid_subregion", lambda z: "NYLI")
    data = export_service.build_carbon_performance_workbook(
        _building(), [7], model_manager=None, cost_calculator=None,
    )
    assert_valid_xlsx(data)


def test_named_ranges_survive(monkeypatch):
    monkeypatch.setattr(export_service, "assess_building_for_export", lambda *a, **k: _fake_result())
    monkeypatch.setattr(export_service, "get_egrid_subregion", lambda z: "NYLI")
    data = export_service.build_carbon_performance_workbook(
        _building(), [7], model_manager=None, cost_calculator=None,
    )
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Export BlueLynx" in wb.sheetnames
    assert "ESPM_Prop_Types" in wb.defined_names      # reference-table range preserved


def test_fallback_to_applicable_when_no_selection(monkeypatch):
    monkeypatch.setattr(export_service, "assess_building_for_export", lambda *a, **k: _fake_result())
    monkeypatch.setattr(export_service, "get_egrid_subregion", lambda z: "NYLI")
    data = export_service.build_carbon_performance_workbook(
        _building(), [], model_manager=None, cost_calculator=None,
    )
    ws = openpyxl.load_workbook(io.BytesIO(data))["Export BlueLynx"]
    assert ws["C6"].value == "LED Lighting"   # applicable measure used as fallback
