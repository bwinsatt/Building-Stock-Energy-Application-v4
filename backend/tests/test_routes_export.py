"""Test the carbon performance export endpoint (service mocked + full e2e)."""

from fastapi.testclient import TestClient

from app.main import app
from app.api import routes
from tests.test_export_service import _fake_result, _building, assert_valid_xlsx


def test_export_endpoint_returns_xlsx(monkeypatch):
    monkeypatch.setattr(
        routes, "build_carbon_performance_workbook", lambda *a, **k: b"PK\x03\x04fake"
    )
    with TestClient(app) as client:
        payload = {
            "building": _building().model_dump(),
            "result": _fake_result().model_dump(),
            "selected_upgrade_ids": [7],
            "espm_property_type": "Office",
        }
        resp = client.post("/export/carbon-performance", json=payload)

    assert resp.status_code == 200
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "CarbonPerformance_00501.xlsx" in resp.headers["content-disposition"]
    assert resp.content == b"PK\x03\x04fake"


def test_export_endpoint_end_to_end(office_input):
    """Inputs -> /assess (runs the model) -> /export (fills the workbook)."""
    with TestClient(app) as client:
        # 1. Run the real assessment.
        assess_resp = client.post(
            "/assess", json={"buildings": [office_input.model_dump()]}
        )
        assert assess_resp.status_code == 200
        result = assess_resp.json()["results"][0]
        assert result.get("rates") is not None      # rates surfaced for export

        # 2. Export using the computed result (no model work on this call).
        export_resp = client.post(
            "/export/carbon-performance",
            json={
                "building": office_input.model_dump(),
                "result": result,
                "selected_upgrade_ids": [],          # exercise the applicable fallback
                "espm_property_type": "Office",
            },
        )

    assert export_resp.status_code == 200
    assert export_resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Full package integrity: no external links, no dangling relationships
    # (this is the corruption Excel flagged but openpyxl tolerates).
    assert_valid_xlsx(export_resp.content)

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(export_resp.content))
    ws = wb["Export BlueLynx"]
    assert "ESPM_Prop_Types" not in wb.defined_names   # external ref stripped, not kept
    assert ws["C6"].value is not None and str(ws["C6"].value).strip() != ""
    assert ws["AJ8"].value == office_input.zipcode
    assert ws["AJ9"].value == "Office"
    assert ws["AJ24"].value == office_input.sqft
    assert ws["AJ16"].value is not None
    assert ws["AJ32"].value == "NREL Cambium, 2022"
