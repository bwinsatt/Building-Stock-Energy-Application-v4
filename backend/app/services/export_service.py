"""Carbon Performance import workbook export.

Fills the SiteLynx Carbon Performance import template ("Export BlueLynx" sheet)
with an already-computed single-building assessment result. No model inference is
performed here, the caller passes the result produced by `/assess`, so the export
is a fast, pure template fill. Only scalar cell values are written; the template's
reference-table named ranges and styles are preserved.
"""

import io
import re
from pathlib import Path

import openpyxl

from app.constants import KWH_TO_KBTU, KWH_PER_THERM, KBTU_PER_THERM
from app.schemas.request import BuildingInput
from app.schemas.response import BuildingResult
from app.services.preprocessor import get_egrid_subregion

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "data" / "templates"
    / "sitelynx_carbon_performance_import.xlsx"
)
SHEET_NAME = "Export BlueLynx"
FIRST_MEASURE_ROW = 6
MAX_MEASURE_ROWS = 30
GRID_DECARB_MODEL = "NREL Cambium, 2022"

# A token like ``[1]`` inside a formula/defined-name addresses an *external*
# workbook (here the SharePoint Carbon Performance calculator). Excel resolves it
# through an externalLink part; with that part gone the reference dangles.
_EXTERNAL_REF = re.compile(r"\[\d+\]")


def _drop_external_references(wb):
    """Remove the external-workbook links and every defined name that points at one.

    The template carries an ``externalLinks`` part (the SharePoint calculator) plus
    ~60 defined names whose formulas reference it as ``[1]…`` (e.g.
    ``[1]Calculations!$D$22``). openpyxl cannot round-trip that part, so we clear it,
    but clearing it alone leaves those defined names referencing a workbook index
    that no longer exists. Excel treats a defined name pointing at a missing external
    book as a corruption and shows the "We found a problem with some content" repair
    prompt, even though the OOXML relationships are otherwise intact. The import sheet
    never needs these names (the real calculator supplies its own), so we strip both
    the link and the names, matching the hand-cleaned template that opens cleanly.
    """
    wb._external_links = []
    scopes = [wb.defined_names] + [ws.defined_names for ws in wb.worksheets]
    for scope in scopes:
        for name in [n for n in scope if _EXTERNAL_REF.search(scope[n].value or "")]:
            del scope[name]


def _select_measures(measures, selected_upgrade_ids):
    """Checked measures, else fall back to applicable individual measures. Capped at 30."""
    selected_ids = set(selected_upgrade_ids or [])
    selected = [m for m in measures if m.upgrade_id in selected_ids]
    if not selected:
        selected = [m for m in measures if m.applicable and m.category != "package"]
    return selected[:MAX_MEASURE_ROWS]


def build_carbon_performance_workbook(
    building: BuildingInput,
    result: BuildingResult,
    selected_upgrade_ids,
    espm_property_type=None,
) -> bytes:
    """Build a filled Carbon Performance import workbook from a precomputed result."""
    sqft = building.sqft
    rates = result.rates
    rate_elec = (rates.electricity if rates else 0.0) or 0.0
    rate_gas = (rates.natural_gas if rates else 0.0) or 0.0

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _drop_external_references(wb)

    ws = wb[SHEET_NAME]

    # --- Measure rows ---
    for i, m in enumerate(_select_measures(result.measures, selected_upgrade_ids)):
        r = FIRST_MEASURE_ROW + i
        ws[f"C{r}"] = m.name
        if m.cost is not None:
            ws[f"D{r}"] = m.cost.installed_cost_total
        if m.electricity_savings_kwh is not None:
            elec_kwh = m.electricity_savings_kwh * sqft
            ws[f"L{r}"] = round(elec_kwh, 0)
            if rate_elec:
                ws[f"H{r}"] = round(elec_kwh * rate_elec, 2)
        if m.gas_savings_therms is not None:
            gas_therms = m.gas_savings_therms * sqft
            ws[f"M{r}"] = round(gas_therms, 1)
            if rate_gas:
                ws[f"I{r}"] = round(gas_therms * KWH_PER_THERM * rate_gas, 2)

    # --- Baseline metrics (column AJ) ---
    ws["AJ8"] = building.zipcode
    if espm_property_type:
        ws["AJ9"] = espm_property_type
    ws["AJ11"] = result.input_summary.state
    egrid = get_egrid_subregion(building.zipcode)
    if egrid:
        ws["AJ14"] = egrid
    ws["AJ15"] = result.input_summary.climate_zone

    elec_kbtu_sf = result.baseline.eui_by_fuel.electricity
    gas_kbtu_sf = result.baseline.eui_by_fuel.natural_gas
    base_elec_kwh = elec_kbtu_sf / KWH_TO_KBTU * sqft
    base_gas_therms = gas_kbtu_sf / KBTU_PER_THERM * sqft

    ws["AJ16"] = round(base_elec_kwh, 0)
    if rate_elec:
        ws["AJ17"] = round(base_elec_kwh * rate_elec, 2)
    ws["AJ20"] = round(base_gas_therms, 1)
    if rate_gas:
        base_gas_kwh = gas_kbtu_sf / KWH_TO_KBTU * sqft
        ws["AJ21"] = round(base_gas_kwh * rate_gas, 2)
    ws["AJ24"] = sqft
    ws["AJ32"] = GRID_DECARB_MODEL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
